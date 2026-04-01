"""
BIST Finansal Tablolar — Masaüstü Uygulaması
=============================================
Çift tıkla aç, hisse kodunu yaz, Excel oluşsun.

Gereksinim: Python 3.8+  (tkinter dahil gelir)
İlk açılışta eksik kütüphaneler otomatik kurulur.
"""

log = lambda *args: open('debug.log', 'a').write(' '.join(map(str, args)) + '\n')

# ══════════════════════════════════════════════════════════════════════════════
# 1.  Python & paket kontrolü  (GUI açılmadan önce)
# ══════════════════════════════════════════════════════════════════════════════
import sys, subprocess
import importlib.util as importlib_util

if sys.version_info < (3, 8):
    import tkinter as tk
    from tkinter import messagebox
    root = tk.Tk(); root.withdraw()
    messagebox.showerror("Hata", f"Python 3.8+ gerekli.\nMevcut: {sys.version}\nhttps://python.org/downloads")
    sys.exit(1)

PAKETLER = [
    ("borsapy",           "borsapy"),
    ("isyatirimhisse",    "isyatirimhisse"),
    ("openpyxl",          "openpyxl"),
    ("pandas",            "pandas"),
    ("requests",          "requests"),
    ("selenium",          "selenium"),
    ("webdriver_manager", "webdriver-manager"),
    ("bs4",               "beautifulsoup4"),
]

eksik = [p for m, p in PAKETLER if not importlib_util.find_spec(m)]
if eksik:
    # Tkinter yükleme ekranı göster
    import tkinter as tk
    yukleme = tk.Tk()
    yukleme.title("Kurulum"); yukleme.geometry("420x120")
    yukleme.resizable(False, False)
    tk.Label(yukleme, text="⏳ Gerekli kütüphaneler kuruluyor...",
             font=("Arial", 11), pady=20).pack()
    lbl = tk.Label(yukleme, text=", ".join(eksik), font=("Arial", 9), fg="#555")
    lbl.pack()
    yukleme.update()
    subprocess.check_call([sys.executable, "-m", "pip", "install",
                           "--quiet", "--upgrade"] + eksik)
    yukleme.destroy()

# ══════════════════════════════════════════════════════════════════════════════
# 2.  Normal import'lar
# ══════════════════════════════════════════════════════════════════════════════
import os, re, time, threading, traceback
import tkinter as tk
import tkinter.scrolledtext as stext
from tkinter import ttk, messagebox, filedialog
from datetime import datetime

import requests as req_lib
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════════
# 3.  Sabitler & Veriler
# ══════════════════════════════════════════════════════════════════════════════
BASLANGIC = 2024
BITIS     = 2025
DONEM_AY  = "12"   # "12","9","6","3","ALL"

SEKME_RENK = {"Bilanço":"1F4E79","Gelir Tablosu":"375623",
              "Nakit Akım":"7B2C2C","Dipnot":"5C3317","Analiz":"6A1B9A"}

ANA_KW = {
    "Bilanço":       ["TOPLAM","DURAN VARLIK","DÖNEN VARLIK","ÖZKAYNAKLAR",
                      "YÜKÜMLÜLÜK","VARLIKLAR","KAYNAKLAR","AKTİF","PASİF"],
    "Gelir Tablosu": ["HASILAT","BRÜT","FAALİYET KÂRI","DÖNEM KÂRI",
                      "VERGİ ÖNCESİ","NET HASILAT","TOPLAM KAPSAMLI"],
    "Nakit Akım":    ["FAALİYETLERİNDEN NAKİT","NAKİT VE NAKİT BENZERLERİ",
                      "NET ARTIŞ","DÖNEM SONU NAKİT","NAKİT AKIŞI"],
    "Dipnot":        ["TOPLAM","NET","GENEL"],
    "Analiz":        ["ORAN","NET BORÇ","FAVÖK"],
}

DONEM_SECENEKLERI = [
    ("Yıllık (12)", "12"),
    ("9 Aylık", "9"),
    ("6 Aylık", "6"),
    ("3 Aylık", "3"),
    ("Tüm Dönemler", "ALL"),
]

# Analiz motoru — Bilanço/Gelir/Nakit satır adları (TFRS + borsapy; Türkçe İ/ı farkına dayanıklı)
ADAY_DONEN_VARLIK = [
    "DÖNEN VARLIKLAR TOPLAMI", "DÖNEN VARLIKLAR", "(ARA TOPLAM)", "ARA TOPLAM",
    "CURRENT ASSETS", "TOTAL CURRENT ASSETS",
    "DÖNEN AKTİFLER", "TOPLAM DÖNEN AKTİFLER", "CARI AKTİFLER", "TOPLAM CARİ AKTİFLER",
    "DÖNEN VARLIKLAR TOPLAMI", "DÖNEN VARLIKLAR TOPLAMI",
]
ADAY_KV_YUK = [
    "KISA VADELİ YÜKÜMLÜLÜKLER TOPLAMI", "KISA VADELİ YÜKÜMLÜLÜKLER",
    "CURRENT LIABILITIES", "TOTAL CURRENT LIABILITIES",
    "KISA VADELİ BORÇLAR", "TOPLAM KISA VADELİ BORÇLAR", "CARI BORÇLAR", "TOPLAM CARİ BORÇLAR",
    "KISA VADELİ YÜKÜMLÜLÜKLER TOPLAMI", "KISA VADELİ YÜKÜMLÜLÜKLER TOPLAMI",
    "BORÇLAR", "TOPLAM BORÇLAR", "KISA VADELİ BORÇLAR TOPLAMI",
]
ADAY_UV_YUK = [
    "UZUN VADELİ YÜKÜMLÜLÜKLER TOPLAMI", "UZUN VADELİ YÜKÜMLÜLÜKLER",
    "NON-CURRENT LIABILITIES", "LONG TERM LIABILITIES", "LONG-TERM LIABILITIES",
]
ADAY_TOPLAM_YUK = ["TOPLAM YÜKÜMLÜLÜKLER", "TOTAL LIABILITIES", "LIABILITIES TOTAL"]
ADAY_OZKAYNAK = [
    "ÖZKAYNAKLAR TOPLAMI", "ÖZKAYNAKLAR", "TOPLAM ÖZKAYNAKLAR",
    "TOTAL EQUITY", "TOTAL SHAREHOLDERS EQUITY", "STOCKHOLDERS EQUITY",
]
ADAY_NAKIT = [
    "NAKİT VE NAKİT BENZERLERİ", "NAKIT VE NAKIT BENZERLERI",
    "CASH AND CASH EQUIVALENTS", "CASH & CASH EQUIVALENTS",
]
ADAY_FIN_BORC_KV = [
    "FİNANSAL BORÇLAR (KV)", "FINANSAL BORÇLAR (KV)", "FINANSAL BORCLAR (KV)",
    "SHORT TERM BORROWINGS", "SHORT-TERM DEBT",
]
ADAY_FIN_BORC_UV = [
    "FİNANSAL BORÇLAR (UV)", "FINANSAL BORÇLAR (UV)",
    "LONG TERM BORROWINGS", "LONG-TERM DEBT",
]
ADAY_HASILAT = [
    "NET HASILAT", "HASILAT", "TOPLAM GELİRLER", "TOPLAM GELIRLER",
    "SATIŞ GELİRLERİ", "SATIS GELIRLERI", "SATIŞLAR", "NET SATIŞLAR",
    "REVENUE", "TOTAL REVENUE", "NET SALES", "SALES",
]
ADAY_BRUT_KAR = [
    "BRÜT KÂR/ZARAR", "BRÜT KAR/ZARAR", "BRÜT KÂR", "BRÜT KAR",
    "BRÜT KAR (ZARAR)", "GROSS PROFIT", "GROSS INCOME",
]
ADAY_FAALIYET_KAR = [
    "ESAS FAALİYET KÂRI/ZARARI", "ESAS FAALİYET KARI/ZARARI",
    "FAALİYET KÂRI/ZARARI", "FAALİYET KARI/ZARARI",
    "FAALİYET KARI (ZARARI)", "OPERATING INCOME", "OPERATING PROFIT", "EBIT",
]
ADAY_NET_KAR = [
    "DÖNEM KÂRI/ZARARI", "DÖNEM KARI/ZARARI", "NET DÖNEM KÂRI",
    "SÜRDÜRÜLEN FAALİYETLER DÖNEM KÂRI",
    "SÜRDÜRÜLEN FAALİYETLER DÖNEM KARI/ZARARI", "DÖNEM KARI (ZARARI)",
    "NET INCOME", "NET PROFIT", "PROFIT FOR THE PERIOD", "NET EARNINGS",
]
ADAY_AMORT = [
    "AMORTİSMAN VE İTFA PAYLARI", "AMORTISMAN VE ITFA PAYLARI", "AMORTİSMAN",
    "DEPRECIATION AND AMORTIZATION", "DEPRECIATION & AMORTIZATION",
    "AMORTISMAN GIDERLERI", "AMORTISMAN GİDERLERİ",
]
ADAY_ISLETME_NAKIT = [
    "İŞLETME FAALİYETLERİNDEN NAKİT AKIŞI", "İŞLETME FAALİYETLERİNDEN NAKİT",
    "OPERATING CASH FLOW", "CASH FROM OPERATIONS", "NET CASH FROM OPERATING",
    "İŞLETME FAALİYETLERİNDEN KAYNAKLANAN NET NAKİT",
]
ADAY_YATIRIM_NAKIT = [
    "YATIRIM FAALİYETLERİNDEN NAKİT AKIŞI", "YATIRIM FAALİYETLERİNDEN NAKİT",
    "INVESTING CASH FLOW", "CASH FROM INVESTING",
    "YATIRIM FAALİYETLERİNDEN KAYNAKLANAN NAKİT",
]
ADAY_TOPLAM_VARLIK = [
    "TOPLAM VARLIKLAR", "AKTİF (VARLIKLAR) TOPLAMI", "VARLIKLAR TOPLAMI",
    "TOTAL ASSETS", "TOTAL ASSET",
]
ADAY_STOK = ["STOKLAR", "STOK", "INVENTORIES", "INVENTORY"]

FINANS_HISSELER = {
    "AKBNK","GARAN","ISCTR","YKBNK","HALKB","VAKBN","QNBFB","ALBRK",
    "ISBTR","SKBNK","TSKB","ICBCT","KLNMA","TKFEN","ALARK",
    "AKGRT","ANSGR","GUSGR","RAYSG","TURSG","ANHYT","AVIVASA",
    "ISMEN","GEDIK","ISFIN","FINBN",
    "EKGYO","ISGYO","TRGYO","ALGYO","OZGYO","VKGYO","PEKGY","SNGYO",
    "KRGYO","DGGYO","MRGYO","HLGYO","NUGYO","RYGYO","ATAGY","AGYO",
    "AVGYO","DZGYO","EGPRO","OZKGY","PEGYO","TSGYO","YGGYO",
    "BIGYO","BYGYO","CYGYO","TACTR","VKFYO","GLBMD",
}

FALLBACK_MAP = {
    "1A":"DÖNEN VARLIKLAR","1AA":"Nakit ve Nakit Benzerleri",
    "1AB":"Finansal Yatırımlar (KV)","1AC":"Ticari Alacaklar",
    "1AD":"İlişkili Taraflardan Ticari Alacaklar",
    "1AE":"Diğer Alacaklar","1AEA":"İlişkili Taraflardan Diğer Alacaklar",
    "1AEB":"İlişkili Olmayan Taraflardan Diğer Alacaklar",
    "1AF":"Türev Araçlar (KV)","1AG":"Stoklar",
    "1AH":"Peşin Ödenmiş Giderler (KV)","1AI":"Cari Dönem Vergi Varlığı",
    "1AJ":"Diğer Dönen Varlıklar","1AK":"Satış Amaçlı Duran Varlıklar",
    "1AL":"DÖNEN VARLIKLAR TOPLAMI",
    "1B":"DURAN VARLIKLAR","1BA":"Finansal Yatırımlar (UV)",
    "1BB":"Özkaynak Yöntemiyle Değerlenen Yatırımlar",
    "1BBA":"Bağlı Ortaklıklar","1BBB":"İş Ortaklıkları",
    "1BC":"Ticari Alacaklar (UV)","1BD":"Diğer Alacaklar (UV)",
    "1BE":"Türev Araçlar (UV)","1BF":"Maddi Duran Varlıklar",
    "1BFA":"Yatırım Amaçlı Gayrimenkuller",
    "1BG":"Kullanım Hakkı Varlıkları","1BH":"Maddi Olmayan Duran Varlıklar",
    "1BHA":"Şerefiye","1BHB":"Diğer Maddi Olmayan Duran Varlıklar",
    "1BI":"Peşin Ödenmiş Giderler (UV)","1BJ":"Ertelenmiş Vergi Varlığı",
    "1BK":"Diğer Duran Varlıklar","1BL":"DURAN VARLIKLAR TOPLAMI",
    "1BM":"AKTİF (VARLIKLAR) TOPLAMI",
    "2A":"KISA VADELİ YÜKÜMLÜLÜKLER","2AA":"Finansal Borçlar (KV)",
    "2AAA":"Banka Kredileri (KV)","2AAB":"Çıkarılmış Tahviller (KV)",
    "2AAC":"Finansal Kiralama Yükümlülükleri (KV)",
    "2AAD":"Diğer Finansal Borçlar (KV)","2AAE":"Kira Yükümlülükleri (KV)",
    "2AAF":"Faktoring Borçları","2AAG":"Diğer Finansal Borçlar (Detay)",
    "2AAGAA":"Repo Yükümlülükleri","2AAGAB":"Para Piyasası Borçları",
    "2AAGAC":"Bankalararası Para Piyasası Borçları",
    "2AAGB":"Müşteri Mevduatları","2AAGC":"Yurt İçi Bankalar Mevduatı",
    "2AAGCA":"Yurt İçi Bankalar Vadesiz Mevduatı",
    "2AAGD":"Yurt Dışı Bankalar Mevduatı",
    "2AAGE":"Kıymetli Maden Depo Hesapları",
    "2AAGF":"Özel Finans Kurumları Mevduatı",
    "2AAGG":"Fon Toplama Hesapları","2AAGH":"Diğer Mevduatlar",
    "2AB":"UV Borcun KV Kısmı","2AC":"Ticari Borçlar",
    "2ACA":"İlişkili Taraflara Ticari Borçlar",
    "2ACB":"İlişkili Olmayan Taraflara Ticari Borçlar",
    "2AD":"Çalışanlara Borçlar","2AE":"Diğer Borçlar",
    "2AEA":"İlişkili Taraflara Diğer Borçlar",
    "2AEB":"İlişkili Olmayan Taraflara Diğer Borçlar",
    "2AF":"Türev Araçlar (KV, Pasif)","2AG":"Ertelenmiş Gelirler (KV)",
    "2AH":"Dönem Kârı Vergi Yükümlülüğü","2AI":"Kısa Vadeli Karşılıklar",
    "2AIA":"Çalışanlara Fayda Karşılıkları (KV)",
    "2AIB":"Diğer Kısa Vadeli Karşılıklar",
    "2AJ":"Diğer KV Yükümlülükler",
    "2AK":"Satış Amaçlı Varlıklara İlişkin Yükümlülükler",
    "2AL":"Kira Yükümlülükleri (KV)","2AM":"KISA VADELİ YÜKÜMLÜLÜKLER TOPLAMI",
    "2B":"UZUN VADELİ YÜKÜMLÜLÜKLER","2BA":"Finansal Borçlar (UV)",
    "2BAA":"Banka Kredileri (UV)","2BAB":"Çıkarılmış Tahviller (UV)",
    "2BAC":"Finansal Kiralama Yükümlülükleri (UV)",
    "2BAD":"Diğer Finansal Borçlar (UV)",
    "2BB":"Ticari Borçlar (UV)","2BBA":"İlişkili Taraflara UV Ticari Borçlar",
    "2BBB":"İlişkili Olmayan Taraflara UV Ticari Borçlar",
    "2BBBA":"UV Ticari Borçlar (Detay)",
    "2BC":"Diğer Borçlar (UV)","2BD":"Türev Araçlar (UV, Pasif)",
    "2BDA":"Döviz Swap Borçları (UV)",
    "2BE":"Ertelenmiş Gelirler (UV)","2BF":"Uzun Vadeli Karşılıklar",
    "2BFA":"Çalışanlara Fayda Karşılıkları (UV)",
    "2BFB":"Diğer Uzun Vadeli Karşılıklar",
    "2BG":"Cari Dönem Vergisi Borçları (UV)",
    "2BH":"Ertelenmiş Vergi Yükümlülüğü",
    "2BI":"Diğer UV Yükümlülükler","2BJ":"Kira Yükümlülükleri (UV)",
    "2BK":"UZUN VADELİ YÜKÜMLÜLÜKLER TOPLAMI",
    "2C":"ÖZKAYNAKLAR","2CA":"Ana Ortaklığa Ait Özkaynaklar",
    "2CAA":"Ödenmiş Sermaye","2CAB":"Geri Alınmış Paylar (-)",
    "2CAC":"Sermaye Düzeltme Farkları",
    "2CAD":"Paylara İlişkin Primler (İskontolar)",
    "2CAE":"Yeniden Sınıflandırılmayacak Birikmiş OKG",
    "2CAF":"Yeniden Sınıflandırılacak Birikmiş OKG",
    "2CAG":"Kârdan Ayrılan Kısıtlanmış Yedekler",
    "2CAH":"Geçmiş Yıllar Kâr/Zararları","2CAI":"Net Dönem Kâr/Zararı",
    "2CB":"Kontrol Gücü Olmayan Paylar",
    "2CC":"ÖZKAYNAKLAR TOPLAMI","2CD":"PASİF (KAYNAKLAR) TOPLAMI",
    "2N":"Sermaye Benzeri Borçlar","2O":"Diğer Yükümlülükler",
    "2OA":"İhraç Edilen Menkul Kıymetler",
    "2OC":"Finansal Borçlar (Diğer)","2OCA":"Diğer Finansal Borçlar A",
    "2OCB":"Diğer Finansal Borçlar B","2OCC":"Diğer Finansal Borçlar C",
    "2OCD":"Diğer Finansal Borçlar D","2OCE":"Diğer Finansal Borçlar E",
    "2OCF":"Diğer Finansal Borçlar F",
    "2OD":"Sigorta Teknik Karşılıkları",
    "2ODA":"Hayat Sigorta Matematik Karşılıkları",
    "2ODB":"Diğer Teknik Karşılıklar",
    "3B":"Birikmiş Diğer Kapsamlı Gelir/Gider",
    "3C":"Yabancı Para Çevrim Farkları",
    "3CAA":"Bağlı Ortaklık YP Çevrim Farkları",
    "3CAB":"İştirak YP Çevrim Farkları",
    "3H":"Kâr veya Zararda Yeniden Sınıflandırılacak OKG",
    "3HA":"Nakit Akış Riskten Korunma Fonu",
    "3HAA":"Nakit Akış Korunması Etkin Kısım",
    "3Z":"Diğer Özkaynak Kalemleri",
    "4B":"Özkaynak Değişim Tablosu Kalemleri",
    "4BA":"Kâr Payı Dağıtımı","4BB":"Sermaye Artırımı",
    "4BC":"Hisse Geri Alımı","4BD":"OKG Değişimi",
    "4CA":"Ana Ortaklık Özkaynakları (Dönem Sonu)",
    "4CAA":"Ödenmiş Sermaye (Dönem Sonu)",
    "4CAK":"Toplam Özkaynaklar (Dönem Sonu)",
    "4CAL":"TOPLAM ÖZKAYNAKLAR",
    "5A":"HASILAT","5AA":"Yurt İçi Satışlar","5AB":"Yurt Dışı Satışlar",
    "5B":"Satış İndirimleri (-)","5C":"NET HASILAT",
    "5D":"Satışların Maliyeti (-)","5E":"BRÜT KÂR/ZARAR",
    "5F":"Genel Yönetim Giderleri (-)","5G":"Pazarlama Giderleri (-)",
    "5H":"Araştırma ve Geliştirme Giderleri (-)",
    "5I":"Esas Faaliyetlerden Diğer Gelirler",
    "5J":"Esas Faaliyetlerden Diğer Giderler (-)",
    "5K":"ESAS FAALİYET KÂRI/ZARARI",
    "5L":"Özkaynak Yöntemiyle Değerlenen Yatırım Kârları",
    "5P":"Finansman Gelirleri","5Q":"Finansman Giderleri (-)",
    "5R":"SÜRDÜRÜLEN FAALİYETLER VERGİ ÖNCESİ KÂRI",
    "5S":"Vergi Geliri/(Gideri)","5SA":"Dönem Vergi Geliri/(Gideri)",
    "5SB":"Ertelenmiş Vergi Geliri/(Gideri)",
    "5T":"SÜRDÜRÜLEN FAALİYETLER DÖNEM KÂRI",
    "5V":"DÖNEM KÂRI/ZARARI",
    "5WA":"Kontrol Gücü Olmayan Paylara Ait Kâr",
    "5WB":"Ana Ortaklık Paylarına Ait Kâr",
    "5X":"HİSSE BAŞINA KAZANÇ",
    "6A":"DİĞER KAPSAMLI GELİR UNSURLARI","6B":"TOPLAM KAPSAMLI GELİR",
    "7A":"İŞLETME FAALİYETLERİNDEN NAKİT AKIŞLARI",
    "7AA":"Dönem Kârı/Zararı","7ABA":"Amortisman ve İtfa Payları",
    "7ABB":"Değer Düşüklüğü Karşılıkları","7ABC":"Karşılıklar",
    "7ABD":"Faiz Gelirleri ve Giderleri","7ABE":"Kur Farkı Zararları",
    "7AC":"İşletme Sermayesindeki Değişimler",
    "7ACA":"Ticari Alacaklardaki Değişim","7ACB":"Stoklardaki Değişim",
    "7ACC":"Ticari Borçlardaki Değişim",
    "7AD":"Ödenen Vergiler","7AE":"İŞLETME FAALİYETLERİNDEN NAKİT AKIŞI",
    "7B":"YATIRIM FAALİYETLERİNDEN NAKİT AKIŞLARI",
    "7BA":"Maddi/Maddi Olmayan Duran Varlık Alımları (-)",
    "7BB":"Maddi/Maddi Olmayan Duran Varlık Satışları",
    "7BC":"Finansal Yatırım Alımları (-)","7BD":"Finansal Yatırım Satışları",
    "7BF":"Alınan Faizler","7BH":"YATIRIM FAALİYETLERİNDEN NAKİT AKIŞI",
    "7C":"FİNANSMAN FAALİYETLERİNDEN NAKİT AKIŞLARI",
    "7CA":"Borçlanmalardan Nakit Girişleri","7CB":"Borç Ödemeleri (-)",
    "7CC":"Finansal Kiralama Ödemeleri (-)","7CD":"Ödenen Faizler (-)",
    "7CE":"Ödenen Kâr Payları (-)","7CF":"Sermaye Artırımından Nakit",
    "7CH":"FİNANSMAN FAALİYETLERİNDEN NAKİT AKIŞI",
    "7D":"KUR FARKLARININ NAKİT ÜZERİNDEKİ ETKİSİ",
    "7E":"NAKİT VE NAKİT BENZERLERİNDEKİ NET ARTIŞ/AZALIŞ",
    "7F":"Dönem Başı Nakit ve Nakit Benzerleri",
    "7G":"DÖNEM SONU NAKİT VE NAKİT BENZERLERİ",
    "8A":"Ek Nakit Akım Bilgileri",
}

# Dipnot / ek tablo kodları için genişletilmiş harita
DIPNOT_MAP = {
    # 3. grup — Özkaynak kalemleri
    "3A":"Ödenmiş Sermaye","3AA":"Sermaye","3AB":"Sermaye Düzeltme Farkları",
    "3AC":"Geri Alınmış Paylar (-)","3AD":"Paylara İlişkin Primler (İskontolar)",
    "3AE":"Kâr veya Zararda Yeniden Sınıflandırılmayacak Birikmiş OKG",
    "3AF":"Kâr veya Zararda Yeniden Sınıflandırılacak Birikmiş OKG",
    "3AG":"Kârdan Ayrılan Kısıtlanmış Yedekler","3AH":"Geçmiş Yıllar Kâr/Zararları",
    "3AI":"Net Dönem Kâr/Zararı","3AJ":"Kontrol Gücü Olmayan Paylar",
    "3B":"Birikmiş Diğer Kapsamlı Gelir/Gider",
    "3BA":"Yeniden Değerleme Artışları/Azalışları",
    "3BB":"Aktüeryal Kazanç/Kayıp","3BC":"Yabancı Para Çevrim Farkları",
    "3BD":"Finansal Varlıkların Gerçeğe Uygun Değer Farkları",
    "3BE":"Nakit Akış Korunması Kazanç/Kayıpları",
    "3BF":"Yabancı Ülkedeki İşletmeye Net Yatırım Korunması",
    "3C":"Yabancı Para Çevrim Farkları",
    "3CA":"Çevrim Farkları","3CAA":"Bağlı Ortaklık YP Çevrim Farkları",
    "3CAB":"İştirak YP Çevrim Farkları",
    "3D":"Finansal Varlık Gerçeğe Uygun Değer Farkları",
    "3DA":"Satılmaya Hazır Finansal Varlık Farkları",
    "3DB":"Özkaynak Yöntemiyle Değerlenen Yatırım Farkları",
    "3DC":"Gerçeğe Uygun Değer Farkı Kâr/Zarara Yansıtılan Finansal Varlıklar",
    "3DD":"Diğer Finansal Varlık Değer Farkları",
    "3DE":"Türev Araç Gerçeğe Uygun Değer Farkları",
    "3E":"Aktüeryal Kazanç/Kayıplar","3EA":"Tanımlanmış Fayda Planı Kazanç/Kayıp",
    "3F":"Değer Artış Fonu","3FA":"Maddi Duran Varlık Yeniden Değerleme Artışı",
    "3FB":"Maddi Olmayan Duran Varlık Yeniden Değerleme Artışı",
    "3G":"Kısıtlanmış Yedekler","3GA":"Yasal Yedekler","3GB":"Statü Yedekleri",
    "3GC":"Olağanüstü Yedekler",
    "3H":"Kâr veya Zararda Yeniden Sınıflandırılacak OKG",
    "3HA":"Nakit Akış Riskten Korunma Fonu",
    "3HAA":"Nakit Akış Korunması Etkin Kısım",
    "3HAB":"Nakit Akış Korunması Etkin Olmayan Kısım",
    "3HB":"Yabancı Para Çevrim Farkı Fonu",
    "3I":"Kısıtlanmamış Yedekler","3IA":"Geçmiş Yıllar Zararı",
    "3IB":"Dağıtılmamış Kârlar",
    "3Z":"Diğer Özkaynak Kalemleri",
    # 4. grup — Özkaynak değişim tablosu
    "4A":"Dönem Başı Bakiyesi","4AA":"Dönem Başı Ana Ortaklık Özkaynakları",
    "4AB":"Dönem Başı Kontrol Gücü Olmayan Paylar",
    "4B":"Özkaynak Değişim Tablosu Kalemleri",
    "4BA":"Kâr Payı Dağıtımı","4BAA":"Nakit Kâr Payı Dağıtımı",
    "4BAB":"Hisse Senedi Kâr Payı","4BAC":"Diğer Kâr Payı Dağıtımı",
    "4BB":"Sermaye Artırımı","4BBA":"Nakit Sermaye Artırımı",
    "4BBB":"Bedelsiz Sermaye Artırımı",
    "4BC":"Hisse Geri Alımı","4BD":"OKG Değişimi",
    "4BDA":"Yabancı Para Çevrim Farkı Değişimi",
    "4BDB":"Gerçeğe Uygun Değer Değişimi",
    "4BDC":"Aktüeryal Kazanç/Kayıp Değişimi",
    "4BE":"Transferler","4BF":"Diğer Özkaynak Değişimleri",
    "4C":"Dönem Sonu Bakiyesi",
    "4CA":"Ana Ortaklık Özkaynakları (Dönem Sonu)",
    "4CAA":"Ödenmiş Sermaye (Dönem Sonu)",
    "4CAB":"Sermaye Düzeltme Farkları (Dönem Sonu)",
    "4CAC":"Paylara İlişkin Primler (Dönem Sonu)",
    "4CAD":"Birikmiş OKG (Dönem Sonu)",
    "4CAE":"Kısıtlanmış Yedekler (Dönem Sonu)",
    "4CAF":"Geçmiş Yıllar Kâr/Zararları (Dönem Sonu)",
    "4CAG":"Net Dönem Kâr/Zararı (Dönem Sonu)",
    "4CAH":"Diğer Özkaynak Kalemleri (Dönem Sonu)",
    "4CAI":"Toplam Ana Ortaklık Özkaynakları",
    "4CAJ":"Kontrol Gücü Olmayan Paylar (Dönem Sonu)",
    "4CAK":"Toplam Özkaynaklar (Dönem Sonu)",
    "4CAL":"TOPLAM ÖZKAYNAKLAR",
    "4CB":"Kontrol Gücü Olmayan Paylar (Dönem Sonu)",
    "4CBB":"Azınlık Payları Dönem Sonu",
    "4CBE":"Kontrol Gücü Olmayan Pay Değişimleri",
    "4CBF":"Diğer Kontrol Gücü Olmayan Pay Değişimleri",
    # Diğer sık görülen dipnot kodları
    "9A":"Satışların Maliyeti Detayı","9B":"Faaliyet Giderleri Detayı",
    "9C":"Diğer Gelirler/Giderler Detayı","9D":"Finansman Gelirleri/Giderleri Detayı",
    "9E":"Vergi Detayı","9F":"Hisse Başına Kazanç Detayı",
    "9G":"İlişkili Taraf İşlemleri","9H":"Taahhütler ve Koşullu Borçlar",
    "9I":"Kıdem Tazminatı Karşılığı","9J":"Sermaye Taahhütleri",
    "9K":"Kira Taahhütleri","9L":"Dava ve İhtilaflar",
    "9M":"Finansal Risk Yönetimi","9N":"Bölüm Bilgileri",
    "9O":"Sonraki Olaylar",
    # THY / havacılık özel
    "3HC":"Nakit Akış Korunma Fonu (THY)","3HB":"YP Çevrim Farkı Fonu (THY)",
    "3CBE":"Diğer Birikmiş OKG",
}

# ══════════════════════════════════════════════════════════════════════════════
# 4.  Scraper Fonksiyonları
# ══════════════════════════════════════════════════════════════════════════════

def tablo_sinifi(kod):
    k = str(kod).strip().upper()
    if re.match(r'^[1-4]', k):  return "Bilanço"
    elif re.match(r'^[56]', k): return "Gelir Tablosu"
    elif re.match(r'^[78]', k): return "Nakit Akım"
    else:                        return "Dipnot"

def tablo_sinifi_gelismis(kod, ad, has_4x=False, bank_3x_gelir=False):
    """
    Finans hisselerinde tablo sınıflandırma:
    - Önce açıklama metninden (ad) anahtar kelime ile sınıflandır
    - Sonra koda göre fallback uygula
    """
    k = str(kod).strip().upper()
    a = str(ad or "").strip().upper()

    if a:
        if any(x in a for x in ["NAKİT AKIŞ", "NAKIT AKIS", "İŞLETME FAALİYETLERİNDEN", "YATIRIM FAALİYETLERİNDEN"]):
            return "Nakit Akım"
        if any(x in a for x in ["HASILAT", "SATIŞ GELİR", "SATIS GELIR", "DÖNEM KÂRI", "DONEM KARI", "VERGİ ÖNCESİ", "BRÜT KÂR", "BRUT KAR"]):
            return "Gelir Tablosu"
        if any(x in a for x in ["VARLIK", "YÜKÜMLÜLÜK", "YUKUMLULUK", "ÖZKAYNAK", "OZKAYNAK", "KAYNAKLAR"]):
            return "Bilanço"

    # Banka formatı: 1-2 bilanço, 3 gelir (5/6/7/8 kodları hiç yok)
    if bank_3x_gelir:
        if re.match(r'^[12]', k): return "Bilanço"
        if re.match(r'^3', k):    return "Gelir Tablosu"

    # Bazı finans/GYO şemalarında 1-2 bilanço, 3 gelir, 4 nakit geliyor (örn: ATAGY)
    if has_4x:
        if re.match(r'^[12]', k): return "Bilanço"
        if re.match(r'^3', k):    return "Gelir Tablosu"
        if re.match(r'^4', k):    return "Nakit Akım"

    # Genel fallback (mevcut davranış)
    return tablo_sinifi(kod)

def aciklama_bul(kod, api_map):
    k = str(kod).strip()
    return api_map.get(k) or FALLBACK_MAP.get(k) or DIPNOT_MAP.get(k) or k

def hisse_turu_belirle(hisse, log):
    if hisse in FINANS_HISSELER:
        log(f"  → {hisse}: Finans/GYO/Banka (listede)")
        return "finans"
    try:
        import borsapy as bp
        df = bp.Ticker(hisse).balance_sheet
        if df is not None and not df.empty:
            log(f"  → {hisse}: Sanayi/Ticaret (borsapy)")
            return "sanayi"
    except Exception:
        pass
    log(f"  → {hisse}: Finans/GYO/Banka (borsapy hata verdi)")
    return "finans"

def api_aciklama_cek(hisse, yil, donem_ay, grup):
    headers = {"User-Agent":"Mozilla/5.0","Referer":"https://www.isyatirim.com.tr/",
               "Accept":"application/json"}
    for url in [
        f"https://www.isyatirim.com.tr/api/data/financials?stockCode={hisse}&year={yil}&period={donem_ay}&financialGroup={grup}",
        f"https://www.isyatirim.com.tr/api/analiz/temelAnaliz/finansalTablo?hisse={hisse}&yil={yil}&donem={donem_ay}&finansalGrup={grup}",
    ]:
        try:
            r = req_lib.get(url, headers=headers, timeout=15)
            if r.status_code != 200: continue
            veri = r.json()
            liste = veri if isinstance(veri, list) else next(
                (veri[k] for k in ["data","items","rows","value"]
                 if isinstance(veri.get(k), list)), [])
            mapping = {}
            for item in liste:
                if not isinstance(item, dict): continue
                kod  = item.get("itemCode") or item.get("FINANCIAL_ITEM_CODE") or item.get("code") or ""
                acik = (item.get("itemDescTr") or item.get("ITEM_DESC_TR") or
                        item.get("descTr") or item.get("description_tr") or
                        item.get("itemDesc") or item.get("description") or "")
                if kod and acik: mapping[str(kod).strip()] = str(acik).strip()
            if mapping: return mapping
        except Exception: continue
    return {}

def api_aciklama_kapsamli(hisse, donem_s_orijinal, aktif_grup):
    birlesik = {}
    for s in donem_s_orijinal:
        try:
            p = str(s).split("/")
            birlesik.update(api_aciklama_cek(hisse, int(p[0]), int(p[1]), aktif_grup))
        except Exception: continue
    if len(birlesik) < 20:
        for g in ["3","2","1"]:
            if g == aktif_grup: continue
            for s in donem_s_orijinal[:2]:
                try:
                    p = str(s).split("/")
                    for k, v in api_aciklama_cek(hisse, int(p[0]), int(p[1]), g).items():
                        if k not in birlesik: birlesik[k] = v
                except Exception: continue
    return birlesik

# ─── borsapy yolu ─────────────────────────────────────────────────────────────
def borsapy_cek(hisse, prop, log):
    try:
        import borsapy as bp
        log(f"  borsapy → {prop}...")
        df = getattr(bp.Ticker(hisse), prop)
        if df is not None and not df.empty:
            log(f"  ✅ {len(df)} kalem")
            return df
        log("  boş geldi"); return None
    except Exception as e:
        log(f"  hata: {e}"); return None

def borsapy_isle(df, baslangic, bitis):
    df = df.copy()
    if df.index.name or not df.columns.empty:
        df = df.reset_index()
        df = df.rename(columns={df.columns[0]:"Kalem"})
    yeni = {}
    for col in df.columns:
        if col == "Kalem": continue
        for y in range(baslangic, bitis+1):
            if str(y) in str(col): yeni[col]=str(y); break
    df = df.rename(columns=yeni)
    yillar = [str(y) for y in range(baslangic, bitis+1)]
    mevcut = [c for c in df.columns if c in yillar] or [c for c in df.columns if c!="Kalem"]
    # Her zaman eskiden yeniye sırala (borsapy bazen ters sıra döndürür)
    mevcut = sorted(set(mevcut), key=lambda s: int(re.search(r'\d{4}', s).group()) if re.search(r'\d{4}', s) else 0)
    df = df[["Kalem"]+mevcut].copy()
    df["Kalem"] = df["Kalem"].astype(str)
    return df, mevcut

def cek_sanayi(hisse, baslangic, bitis, log):
    tablolar = {}
    for sekme, prop in [("Bilanço","balance_sheet"),
                        ("Gelir Tablosu","income_stmt"),
                        ("Nakit Akım","cashflow")]:
        log(f"\n▶ {sekme} çekiliyor...")
        df_raw = borsapy_cek(hisse, prop, log)
        if df_raw is None: continue
        df, donem = borsapy_isle(df_raw, baslangic, bitis)
        if df.empty or not donem:
            log(f"  ⚠️  {baslangic}-{bitis} dönemi bulunamadı"); continue
        log(f"  Dönem sütunları: {donem}")
        # İlk 2 satırı logla — gerçek değerleri görmek için
        for _, row in df.head(2).iterrows():
            log(f"  ÖRNEK: {row['Kalem'][:30]} | {' | '.join(str(row.get(d,'?')) for d in donem)}")
        tablolar[sekme] = (df, donem)
    return tablolar

# ─── isyatirimhisse yolu ───────────────────────────────────────────────────────
def cek_finans(hisse, baslangic, bitis, log, donem_ay="12"):
    from isyatirimhisse import fetch_financials
    yillar = [str(y) for y in range(baslangic, bitis+1)]
    en_iyi = None  # {"grup","df_ham","donem","score","detay"}

    for grup in ["3","2","1"]:
        try:
            log(f"  Grup {grup} deneniyor...")
            df = fetch_financials(symbols=hisse, start_year=baslangic,
                                  end_year=bitis, exchange="TRY", financial_group=grup)
            if df is None or df.empty: log("  boş"); continue
            log(f"  ✅ {len(df)} satır | Sütunlar: {list(df.columns)}")

            kod_sutunu = next(
                (s for s in df.columns if any(k in str(s).lower()
                 for k in ["item_code","itemcode"]) or str(s).lower()=="code"),
                df.columns[0])

            # Açıklama sütununu bul
            aciklama_sutunu = None
            for s in df.columns:
                sl = str(s).lower()
                if any(k in sl for k in ["desctr","desc_tr","description_tr",
                                          "itemdesctr","item_desc_tr",
                                          "name_tr","financial_item_name_tr","item_name_tr"]):
                    aciklama_sutunu = s; break
            if not aciklama_sutunu:
                for s in df.columns:
                    if s in {kod_sutunu,"stockCode","symbol","HISSE_KODU",
                             "period","financialGroup","financial_group"}: continue
                    sample = df[s].dropna().astype(str)
                    if sample.empty: continue
                    if sample.str.len().mean() > 5 and sample.str.match(r"^-?\d[\d.,]*$").mean() < 0.3:
                        aciklama_sutunu = s; break

            if aciklama_sutunu:
                log(f"  Açıklama sütunu: '{aciklama_sutunu}' ✅")

            gereksiz = {"stockCode","symbol","HISSE_KODU","period",
                        "financialGroup","financial_group"}
            if aciklama_sutunu: gereksiz.add(aciklama_sutunu)
            gereksiz.add(kod_sutunu)
            donem_s = [s for s in df.columns if s not in gereksiz]

            hedef_ay = str(donem_ay or "12").upper()
            if hedef_ay in {"3","6","9","12"}:
                odak = [s for s in donem_s if any(f"{y}/{hedef_ay}" in str(s) for y in yillar)]
                donem_s = sorted(odak) if odak else [s for s in donem_s if any(y in str(s) for y in yillar)] or donem_s
            elif hedef_ay == "ALL":
                donem_s = [s for s in donem_s if any(y in str(s) for y in yillar)] or donem_s
            else:
                yillik = [s for s in donem_s if any(f"{y}/12" in str(s) for y in yillar)]
                donem_s = sorted(yillik) if len(yillik)>=2 else \
                          [s for s in donem_s if any(y in str(s) for y in yillar)] or donem_s

            sutunlar = [kod_sutunu]
            if aciklama_sutunu: sutunlar.append(aciklama_sutunu)
            sutunlar += donem_s

            df_t = df[sutunlar].copy().rename(columns={kod_sutunu:"KOD"})
            if aciklama_sutunu: df_t = df_t.rename(columns={aciklama_sutunu:"ACIKLAMA"})

            # Dönem sütunlarını "YIL[/AY]" -> "YIL_AY" (veya "YIL") standardına getir
            yeniden = {}
            for s in donem_s:
                s_str = str(s)
                m = re.search(r"(20\d{2})(?:/(\d{1,2}))?", s_str)
                if m:
                    yy = m.group(1)
                    aa = m.group(2)
                    if aa and str(donem_ay).upper() == "ALL":
                        yeniden[s] = f"{yy}_{aa.zfill(2)}"
                    elif aa and str(donem_ay).upper() in {"3", "6", "9"}:
                        yeniden[s] = f"{yy}_{str(donem_ay).zfill(2)}"
                    else:
                        yeniden[s] = yy
                else:
                    yeniden[s] = s_str

            donem_s_orijinal = list(yeniden.keys()) if yeniden else donem_s[:]

            if yeniden:
                df_t = df_t.rename(columns=yeniden)
                donem_s = [yeniden.get(s, s) for s in donem_s]

            log(f"  Dönem sütunları: {donem_s}")

            donem_s_orijinal = list(yeniden.keys()) if yeniden else donem_s
            api_map = api_aciklama_kapsamli(hisse, donem_s_orijinal, grup)
            log(f"  API map: {len(api_map)} kod + Fallback: {len(FALLBACK_MAP)} kod")

            def kalem_adi(satir):
                if "ACIKLAMA" in satir.index:
                    v = str(satir["ACIKLAMA"]).strip()
                    if v and v not in ("nan","None",""): return v
                ad = aciklama_bul(satir["KOD"], api_map)
                # API bazı kalemlerde KV/UV ayrımını yazmayabiliyor (örn: "Diğer Borçlar")
                try:
                    kod = str(satir["KOD"]).strip().upper()
                    ad_u = str(ad).strip().upper()
                    if "DİĞER BORÇLAR" in ad_u or "DIĞER BORÇLAR" in ad_u:
                        if "(KV" not in ad_u and "(UV" not in ad_u:
                            if kod.startswith("2A"): ad = f"{ad} (KV)"
                            elif kod.startswith("2B"): ad = f"{ad} (UV)"
                except Exception:
                    pass
                return ad

            df_t["Kalem"]  = df_t.apply(kalem_adi, axis=1)
            kod_seri = df_t["KOD"].astype(str).str.upper().str.strip()
            has_4x = kod_seri.str.match(r"^4").any()
            has_5_8 = kod_seri.str.match(r"^[5678]").any()
            has_3x = kod_seri.str.match(r"^3").any()
            has_1_2 = kod_seri.str.match(r"^[12]").any()
            bank_3x_gelir = bool(has_3x and has_1_2 and not has_5_8)
            df_t["_sinif"] = df_t.apply(
                lambda r: tablo_sinifi_gelismis(
                    r["KOD"], r.get("Kalem", ""),
                    has_4x=has_4x, bank_3x_gelir=bank_3x_gelir
                ),
                axis=1
            )

            dagilim = df_t["_sinif"].value_counts().to_dict()
            log(f"  Tablo dağılımı: {dagilim}")
            adetler = {k: int((df_t["_sinif"] == k).sum()) for k in ["Bilanço","Gelir Tablosu","Nakit Akım"]}
            kapsama = sum(1 for k in adetler.values() if k > 0)
            score = (kapsama * 10_000) + sum(adetler.values())
            if en_iyi is None or score > en_iyi["score"]:
                en_iyi = {"grup": grup, "df_ham": df_t, "donem": donem_s, "score": score, "detay": adetler}
        except ValueError as e:
            log(f"  atlandı: {str(e)[:60]}")
        except Exception as e:
            log(f"  hata: {e}")

    if en_iyi is None:
        return {}
    log(f"  ▶ Seçilen grup: {en_iyi['grup']} | kapsam: {en_iyi['detay']}")
    df_ham, donem_sutunlar = en_iyi["df_ham"], en_iyi["donem"]

    tablolar = {}
    # "Dipnot" sınıfına düşen satırlar aslında tanımlanamayan bilanço kalemleridir,
    # gerçek dipnot Selenium ile ayrıca çekilir — burada Dipnot yazılmıyor
    for sekme in ["Bilanço","Gelir Tablosu","Nakit Akım"]:
        alt = df_ham[df_ham["_sinif"]==sekme][["Kalem"]+donem_sutunlar].copy().reset_index(drop=True)
        if not alt.empty:
            tablolar[sekme] = (alt, donem_sutunlar)
            log(f"  ✅ {sekme}: {len(alt)} kalem")

    # Banka formatında (3x gelir) bazı kaynaklarda nakit akım hiç gelmeyebilir.
    # Kullanıcıya boş geçmek yerine açıklayıcı bir sekme üret.
    kod_seri_all = df_ham["KOD"].astype(str).str.upper().str.strip() if "KOD" in df_ham.columns else pd.Series(dtype=str)
    bank_format = bool(kod_seri_all.str.match(r"^3").any() and kod_seri_all.str.match(r"^[12]").any() and not kod_seri_all.str.match(r"^[5678]").any())
    tablolar["bank_format"] = bank_format
    if bank_format and "Nakit Akım" not in tablolar and donem_sutunlar:
        satir = {"Kalem": "Kaynak sağlayıcı bu banka hissesi için Nakit Akım tablosunu döndürmedi."}
        for d in donem_sutunlar:
            satir[d] = None
        tablolar["Nakit Akım"] = (pd.DataFrame([satir]), donem_sutunlar)
        log("  ℹ️  Nakit Akım: kaynakta bulunamadı, bilgilendirme satırı eklendi.")
    return tablolar

# ─── Dipnot çekici ────────────────────────────────────────────────────────────
def dipnot_cek(hisse, baslangic, bitis, log, donem_ay="12"):
    """
    Dipnot verilerini isyatirimhisse API üzerinden çeker.
    API financial_group=2 dipnot/ek finansal bilgileri içerir.
    Başarısız olursa Selenium ile dener.
    """
    # ── 1. isyatirimhisse ile dene ──────────────────────────────────────────
    try:
        from isyatirimhisse import fetch_financials
        yillar = [str(y) for y in range(baslangic, bitis+1)]
        log("  isyatirimhisse dipnot çekiliyor...")

        for grup in ["1","2","3"]:
            try:
                df = fetch_financials(symbols=hisse, start_year=baslangic,
                                      end_year=bitis, exchange="TRY",
                                      financial_group=grup)
                if df is None or df.empty: continue
                log(f"  Dipnot grup={grup}: {len(df)} satır")

                log(f"  Sütunlar: {list(df.columns)}")

                # Kod sütununu bul
                kod_kol = next((c for c in df.columns
                                if any(k in str(c).lower()
                                       for k in ["item_code","itemcode","code"])),
                               df.columns[0])
                # Açıklama sütununu bul — geniş arama
                acik_kol = None
                for c in df.columns:
                    cl = str(c).lower()
                    if any(k in cl for k in ["desctr","desc_tr","description_tr",
                                              "itemdesctr","item_desc_tr",
                                              "aciklama","tanim","label","name","title"]):
                        acik_kol = c; break
                # Hâlâ bulunamadıysa sayısal olmayan ilk sütunu dene
                if not acik_kol:
                    for c in df.columns:
                        if c == kod_kol: continue
                        if str(c).lower() in {"stockcode","symbol","period",
                                               "financialgroup","financial_group"}: continue
                        sample = df[c].dropna().astype(str)
                        if len(sample) == 0: continue
                        sayi_orani = sample.str.match(r'^-?[\d.,\(\)]+$').mean()
                        if sayi_orani < 0.4 and sample.str.len().mean() > 4:
                            acik_kol = c
                            break
                log(f"  Kod sütunu: '{kod_kol}' | Açıklama sütunu: '{acik_kol}'")

                # Dönem sütunlarını bul ve temizle
                gereksiz = {"stockCode","symbol","HISSE_KODU","period",
                            "financialGroup","financial_group",
                            kod_kol, acik_kol}
                donem_cols = [c for c in df.columns if c not in gereksiz and c]

                hedef_ay = str(donem_ay or "12").upper()
                secili = []
                for c in donem_cols:
                    s = str(c)
                    if not any(y in s for y in yillar):
                        continue
                    if hedef_ay in {"3", "6", "9", "12"}:
                        if any(f"{y}/{hedef_ay}" in s for y in yillar):
                            secili.append(c)
                    else:
                        secili.append(c)
                if not secili:
                    continue
                secili = sorted(secili, key=lambda x: str(x))
                rename_donem = {}
                for c in secili:
                    m = re.search(r"(20\d{2})(?:/(\d{1,2}))?", str(c))
                    if not m:
                        rename_donem[c] = str(c)
                        continue
                    yy = m.group(1)
                    aa = m.group(2)
                    if aa and hedef_ay == "ALL":
                        rename_donem[c] = f"{yy}_{aa.zfill(2)}"
                    elif aa and hedef_ay in {"3", "6", "9"}:
                        rename_donem[c] = f"{yy}_{hedef_ay.zfill(2)}"
                    else:
                        rename_donem[c] = yy

                # Önce sadece gerekli sütunları al
                sutun_listesi = [kod_kol]
                if acik_kol: sutun_listesi.append(acik_kol)
                sutun_listesi += secili

                alt = df[sutun_listesi].copy()
                # Dönem sütunlarını yıla rename et
                alt = alt.rename(columns=rename_donem)
                # Açıklama ve kod sütunlarını standartlaştır
                if acik_kol: alt = alt.rename(columns={acik_kol: "_ACIK"})
                alt = alt.rename(columns={kod_kol: "KOD"})
                rename = rename_donem  # api_map için

                # Kalem adı — tüm gruplardan API map'i topla
                api_map = {}
                for g2 in ["1","2","3"]:
                    try:
                        api_map.update(api_aciklama_kapsamli(hisse, list(rename.keys())[:2], g2))
                    except Exception:
                        pass

                def dp_kalem(row):
                    # 1. Açıklama sütunu varsa kullan
                    if "_ACIK" in row.index:
                        v = str(row["_ACIK"]).strip()
                        if v and v not in ("nan","None","","0"): return v
                    kod = str(row["KOD"]).strip()
                    # 2. API map → FALLBACK_MAP → DIPNOT_MAP → kodu olduğu gibi bırak
                    return (api_map.get(kod)
                            or FALLBACK_MAP.get(kod)
                            or DIPNOT_MAP.get(kod)
                            or kod)
                alt["Kalem"] = alt.apply(dp_kalem, axis=1)

                donem_sutunlar = [rename_donem.get(c, str(c)) for c in secili]
                alt = alt[["Kalem"] + donem_sutunlar].copy().reset_index(drop=True)

                # Boş/anlamsız satırları temizle
                alt = alt[alt["Kalem"].astype(str).str.strip().ne("")]
                alt = alt[alt["Kalem"].astype(str).str.len() > 1]
                alt = alt.reset_index(drop=True)

                if len(alt) > 3:
                    log(f"  ✅ Dipnot: {len(alt)} kalem (isyatirimhisse)")
                    # İlk 3 satırı logla
                    for _, row in alt.head(3).iterrows():
                        vals = ' | '.join(str(row.get(d,'?')) for d in donem_sutunlar)
                        log(f"  DIPNOT ÖRNEK: {str(row.get('Kalem','?'))[:35]} | {vals}")
                    return alt, donem_sutunlar
            except Exception as ex:
                log(f"  Dipnot grup {grup} hata: {str(ex)[:80]}")
                continue
    except ImportError:
        pass
    except Exception as ex:
        log(f"  isyatirimhisse dipnot genel hata: {str(ex)[:80]}")

    # ── 2. Selenium fallback ────────────────────────────────────────────────
    log("  Selenium dipnot fallback deneniyor...")
    df_sel = _selenium_dipnot_cek(hisse, baslangic, bitis, log)
    if df_sel is not None and not df_sel.empty:
        yillar = [str(y) for y in range(baslangic, bitis+1)]
        donem = [c for c in df_sel.columns if c != "Kalem"
                 and any(y in str(c) for y in yillar)]
        if not donem:
            donem = [c for c in df_sel.columns if c != "Kalem"]
        return df_sel, donem

    log("  ⚠️  Dipnot çekilemedi")
    return None, []


def _selenium_dipnot_cek(hisse, baslangic, bitis, log):
    """Selenium ile İş Yatırım dipnot sayfasından veri çeker (fallback)."""
    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.chrome.service import Service
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from webdriver_manager.chrome import ChromeDriverManager
        from bs4 import BeautifulSoup

        url = (f"https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/"
               f"sirket-karti.aspx?hisse={hisse}#tab-3")
        log(f"  Selenium: {url}")

        opts = Options()
        opts.add_argument("--headless=new")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-dev-shm-usage")
        opts.add_argument("--disable-gpu")
        opts.add_argument("--window-size=1920,1080")
        opts.add_argument("--disable-blink-features=AutomationControlled")
        opts.add_experimental_option("excludeSwitches", ["enable-automation"])
        opts.add_argument(
            "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36")
        try:
            driver = webdriver.Chrome(
                service=Service(ChromeDriverManager().install()), options=opts)
        except Exception as e:
            log(f"  ChromeDriver kurulamadı: {e}"); return None

        try:
            driver.get(url)
            time.sleep(10)
            try:
                WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.TAG_NAME, "table")))
                time.sleep(3)
            except Exception:
                pass

            soup = BeautifulSoup(driver.page_source, "html.parser")
            tablolar_html = soup.find_all("table")
            if not tablolar_html: return None

            # En geniş tabloyu seç ama bilanço benzeri kalemler içermeyeni tercih et
            bilanco_kw = {"DÖNEN VARLIKLAR","TOPLAM VARLIKLAR","ÖZKAYNAKLAR",
                          "DURAN VARLIKLAR","KAYNAKLAR"}
            en_genis = max(
                [t for t in tablolar_html if len(t.find_all("tr")) >= 3],
                key=lambda t: len(t.find_all("tr")),
                default=tablolar_html[0]
            )

            tum_satirlar = en_genis.find_all("tr")
            basliklar, satirlar = [], []
            for tr in tum_satirlar:
                hucre_th = tr.find_all("th")
                hucre_td = tr.find_all("td")
                hucre = hucre_th or hucre_td
                huc = [h.get_text(strip=True) for h in hucre]
                if not basliklar and (hucre_th or not satirlar):
                    basliklar = huc
                elif any(h.strip() for h in huc):
                    satirlar.append(huc)

            if not basliklar and satirlar:
                basliklar = satirlar.pop(0)
            if not basliklar or not satirlar: return None

            max_k = max(len(basliklar), max((len(r) for r in satirlar), default=0))
            basliklar += [""] * (max_k - len(basliklar))
            for s in satirlar: s += [""] * (max_k - len(s))

            # Yinelenen başlıkları düzelt
            seen = {}
            clean = []
            for b in basliklar:
                key = b if b else "_bos"
                seen[key] = seen.get(key, 0) + 1
                clean.append(f"{b}_{seen[key]}" if seen[key] > 1 else b)

            df = pd.DataFrame(satirlar, columns=clean)
            df = df.rename(columns={clean[0]: "Kalem"})
            df = df[df["Kalem"].astype(str).str.strip().ne("")]
            df = df.reset_index(drop=True)

            yillar = [str(y) for y in range(baslangic, bitis+1)]
            diger = [c for c in df.columns[1:]
                     if any(y in str(c) for y in yillar)]
            if not diger: diger = list(df.columns[1:])
            df = df[["Kalem"] + diger]
            log(f"  ✅ Selenium dipnot: {len(df)} kalem")
            return df
        finally:
            try: driver.quit()
            except Exception: pass
    except ImportError:
        log("  Selenium kurulu değil"); return None
    except Exception as e:
        log(f"  Selenium hata: {e}"); return None
# ─── Hesap & Excel ────────────────────────────────────────────────────────────
def sayiya_cevir_float(val):
    if val is None or str(val).strip() in ("","-","nan","None","NaN"): return None
    s = str(val).strip()
    # Parantez içindeki negatif sayılar: (1.234.567) → -1234567
    negativ = s.startswith("(") and s.endswith(")")
    if negativ: s = s[1:-1]
    # Türkçe format: binlik ayraç nokta, ondalık virgül → 1.234.567,89
    # İngilizce format: binlik ayraç virgül, ondalık nokta → 1,234,567.89
    s = s.replace(" ", "")
    # Virgül VE nokta varsa: son ayracı ondalık kabul et
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            # Türkçe: 1.234,56 → 1234.56
            s = s.replace(".", "").replace(",", ".")
        else:
            # İngilizce: 1,234.56 → 1234.56
            s = s.replace(",", "")
    elif "," in s:
        # Sadece virgül: Türkçe ondalık 1234,56 → 1234.56
        # Ama 1,234,567 gibi ingilizce binlik de olabilir
        parts = s.split(",")
        if len(parts) > 1 and all(len(p) == 3 for p in parts[1:]):
            s = s.replace(",", "")  # binlik ayraç
        else:
            s = s.replace(",", ".")  # ondalık
    elif "." in s:
        # Sadece nokta: binlik mi ondalık mı?
        parts = s.split(".")
        if len(parts) > 2 or (len(parts) == 2 and len(parts[-1]) == 3 and len(parts[0]) > 0):
            # 1.234.567 veya 1.234 (3 haneli) → binlik ayraç
            s = s.replace(".", "")
        # else: 1.5 gibi ondalık → değiştirme
    try:
        result = float(s)
        return -result if negativ else result
    except:
        return None

def sayiya_cevir_int(val):
    f = sayiya_cevir_float(val)
    if f is None: return None
    try: return int(f)
    except: return val

def _metin_norm(s):
    return re.sub(r"\s+"," ",str(s or "").strip().upper())

def _fold_key(s):
    """
    Türkçe/İngilizce kalem adlarını karşılaştırma için tek forma indirir
    (VADELİ / VADELI, İ/ I farkları dahil).
    """
    t = _metin_norm(s)
    for a, b in [
        ("İ", "I"), ("I", "I"), ("Ş", "S"), ("Ğ", "G"), ("Ü", "U"), ("Ö", "O"), ("Ç", "C"),
        ("Â", "A"), ("Ê", "E"), ("Û", "U"),
    ]:
        t = t.replace(a, b)
    return t

def _yil_sutunu(df, yil):
    """Sütun adı bazen int (2024) bazen str ('2024') olabiliyor."""
    if yil in df.columns:
        return yil
    ys = str(yil)
    if ys in df.columns:
        return ys
    return None

def _satir_deger_bul(df, yil, adaylar):
    """
    df: 'Kalem' + yıl sütunları içeren tablo
    adaylar: tam eşleşme veya substring adayları (Türkçe/İngilizce; borsapy satır adlarıyla uyumlu)
    """
    if df is None or df.empty or "Kalem" not in df.columns:
        return None
    ycol = _yil_sutunu(df, yil)
    if ycol is None:
        return None
    kalemler_fold = df["Kalem"].astype(str).map(_fold_key)
    for a in (adaylar or []):
        ad = _fold_key(a)
        if not ad:
            continue
        st = kalemler_fold.str.strip()
        ix = kalemler_fold[st == ad].index
        if len(ix) == 0:
            ix = kalemler_fold[st.str.startswith(ad + " ") | (st == ad)].index
        if len(ix) == 0:
            mask = kalemler_fold.str.contains(re.escape(ad), na=False)
            not_diger = ~st.fillna("").str.match(r"^(DIGER|DIĞER|DIER|DİĞER)\s", na=False)
            ix = kalemler_fold[mask & not_diger].index
        if len(ix) == 0:
            ix = kalemler_fold[kalemler_fold.str.contains(re.escape(ad), na=False)].index
        if len(ix):
            v = sayiya_cevir_float(df.loc[ix[0], ycol])
            if v is not None:
                return v
    return None

def _satir_deger_satirlari_toplam(df, yil, adaylar):
    """Adaylardan biriyle eşleşen tüm satırların toplamı (örn. iki 'Finansal Borçlar' satırı)."""
    if df is None or df.empty or "Kalem" not in df.columns:
        return None
    ycol = _yil_sutunu(df, yil)
    if ycol is None:
        return None
    kalemler_fold = df["Kalem"].astype(str).map(_fold_key)
    idx_set = set()
    for a in (adaylar or []):
        ad = _fold_key(a)
        if len(ad) < 3:
            continue
        mask = kalemler_fold.str.contains(re.escape(ad), na=False)
        idx_set.update(df.index[mask].tolist())
    toplam = 0.0
    for idx in idx_set:
        v = sayiya_cevir_float(df.loc[idx, ycol])
        if v is not None:
            toplam += v
    return toplam if idx_set else None

def analiz_df_olustur(tablolar, yil_eski, yil_yeni):
    bilanco = tablolar.get("Bilanço", (None, None))[0] if tablolar else None
    gelir   = tablolar.get("Gelir Tablosu", (None, None))[0] if tablolar else None
    nakit   = tablolar.get("Nakit Akım", (None, None))[0] if tablolar else None
    bank_format = tablolar.get("bank_format", False)

    def oranlar(yil):
        donen = _satir_deger_bul(bilanco, yil, ADAY_DONEN_VARLIK)
        kv_yuk = _satir_deger_bul(bilanco, yil, ADAY_KV_YUK)
        uv_yuk = _satir_deger_bul(bilanco, yil, ADAY_UV_YUK)
        ozkay  = _satir_deger_bul(bilanco, yil, ADAY_OZKAYNAK)

        fin_kv = _satir_deger_bul(bilanco, yil, ADAY_FIN_BORC_KV)
        fin_uv = _satir_deger_bul(bilanco, yil, ADAY_FIN_BORC_UV)
        nakit_ = _satir_deger_bul(bilanco, yil, ADAY_NAKIT)

        ebit = _satir_deger_bul(gelir, yil, ADAY_FAALIYET_KAR)
        amort = _satir_deger_bul(nakit, yil, ADAY_AMORT)
        if amort is None:
            amort = _satir_deger_bul(nakit, yil, ["AMORTISMAN GIDER", "AMORTISMAN GİDER", "İTFA PAYI", "ITFA PAYI"])
        favok = (ebit + amort) if (ebit is not None and amort is not None) else None

        cari_oran = (donen / kv_yuk) if (donen is not None and kv_yuk not in (None, 0)) else None
        likit = None
        kisa_borc = None
        if cari_oran is None and bank_format:
            # Banka için alternatif cari oran: likit varlıklar / kısa vadeli borçlar
            likit_aday = ["NAKİT VE NAKİT BENZERLERİ", "MENKUL KIYMETLER", "HAZIR DEĞERLER", "LIKİT VARLIKLAR", "DÖNEN VARLIKLAR"]
            kisa_borc_aday = ["KISA VADELİ BORÇLAR", "MEVDUATLAR", "KISA VADELİ YÜKÜMLÜLÜKLER", "KISA VADELİ BORÇLAR TOPLAMI"]
            likit = _satir_deger_bul(bilanco, yil, likit_aday)
            kisa_borc = _satir_deger_bul(bilanco, yil, kisa_borc_aday)
            if likit is not None and kisa_borc not in (None, 0):
                cari_oran = likit / kisa_borc
        log(f"  {yil} - donen: {donen}, kv_yuk: {kv_yuk}, likit: {likit}, kisa_borc: {kisa_borc}, cari_oran: {cari_oran}")
        yuk_toplam = None
        if kv_yuk is not None and uv_yuk is not None:
            yuk_toplam = kv_yuk + uv_yuk
        elif bilanco is not None:
            yuk_toplam = _satir_deger_bul(bilanco, yil, ADAY_TOPLAM_YUK)
        borc_ozk = (yuk_toplam / ozkay) if (yuk_toplam is not None and ozkay not in (None, 0)) else None

        toplam_fin_borc = _satir_deger_satirlari_toplam(bilanco, yil, ["FINANSAL BORCLAR", "FINANSAL BORÇLAR"])
        if toplam_fin_borc is None:
            if fin_kv is not None and fin_uv is not None:
                toplam_fin_borc = fin_kv + fin_uv
            elif fin_kv is not None:
                toplam_fin_borc = fin_kv
            elif fin_uv is not None:
                toplam_fin_borc = fin_uv

        net_borc = (toplam_fin_borc - nakit_) if (toplam_fin_borc is not None and nakit_ is not None) else None
        net_borc_favok = (net_borc / favok) if (net_borc is not None and favok not in (None, 0)) else None

        return {
            "Cari Oran": cari_oran,
            "Borç/Özkaynak": borc_ozk,
            "Net Borç": net_borc,
            "FAVÖK (≈)": favok,
            "Net Borç/FAVÖK": net_borc_favok,
        }

    o_eski = oranlar(yil_eski)
    o_yeni = oranlar(yil_yeni)

    satirlar = []
    for k in ["Cari Oran","Borç/Özkaynak","Net Borç","FAVÖK (≈)","Net Borç/FAVÖK"]:
        e = o_eski.get(k); y = o_yeni.get(k)
        # =(C-B)/ABS(B) → B=eski(2024), C=yeni(2025) → (yeni-eski)/abs(eski)
        deg = round((y - e) / abs(e) * 100, 2) if (e is not None and y is not None and e != 0) else None
        satirlar.append({"Kalem": k, yil_eski: e, yil_yeni: y, "Değişim %": deg})

    df = pd.DataFrame(satirlar)
    # Sütun sırası garantile
    for col in [yil_eski, yil_yeni, "Değişim %"]:
        if col not in df.columns:
            df[col] = None
    df = df[["Kalem", yil_eski, yil_yeni, "Değişim %"]]
    return df


def yerlesik_ai_analiz(tablolar, yil_eski, yil_yeni, log):
    """
    Tamamen yerleşik kural tabanlı finansal analiz motoru.
    Hiçbir dış bağımlılık yok — internet, API veya kurulum gerektirmez.
    Verileri okur, oranları hesaplar, eşik değerlere göre yorum üretir.
    """
    log("  Yerleşik AI analiz motoru çalışıyor...")

    bilanco = tablolar.get("Bilanço", (None, None))[0]
    gelir   = tablolar.get("Gelir Tablosu", (None, None))[0]
    nakit   = tablolar.get("Nakit Akım", (None, None))[0]

    def v(df, yil, adaylar):
        return _satir_deger_bul(df, yil, adaylar)

    def degisim(eski, yeni):
        if eski is None or yeni is None or eski == 0:
            return None
        return (yeni - eski) / abs(eski) * 100

    def fmt_para(x):
        if x is None: return "veri yok"
        milyar = x / 1_000_000_000
        if abs(milyar) >= 1:
            return f"{milyar:,.2f} Milyar ₺"
        milyon = x / 1_000_000
        return f"{milyon:,.1f} Milyon ₺"

    def fmt_oran(x, ondalik=2):
        if x is None: return "veri yok"
        return f"{x:.{ondalik}f}x"

    def fmt_pct(x):
        if x is None: return "veri yok"
        return f"%{x:+.1f}"

    # ── Tüm verileri her iki dönem için çek ──────────────────────────────────
    def veri_seti(yil):
        donen      = v(bilanco, yil, ADAY_DONEN_VARLIK)
        kv_yuk     = v(bilanco, yil, ADAY_KV_YUK)
        uv_yuk     = v(bilanco, yil, ADAY_UV_YUK)
        toplam_var = v(bilanco, yil, ADAY_TOPLAM_VARLIK)
        ozkay      = v(bilanco, yil, ADAY_OZKAYNAK)
        nakit_     = v(bilanco, yil, ADAY_NAKIT)
        stok       = v(bilanco, yil, ADAY_STOK)
        fin_kv     = v(bilanco, yil, ADAY_FIN_BORC_KV)
        fin_uv     = v(bilanco, yil, ADAY_FIN_BORC_UV)
        hasilat    = v(gelir,   yil, ADAY_HASILAT)
        brut_kar   = v(gelir,   yil, ADAY_BRUT_KAR)
        faaliyet_k = v(gelir,   yil, ADAY_FAALIYET_KAR)
        net_kar    = v(gelir,   yil, ADAY_NET_KAR)
        amort      = v(nakit,   yil, ADAY_AMORT)
        if amort is None:
            amort = v(nakit, yil, ["AMORTISMAN GIDER", "AMORTISMAN GİDER", "İTFA PAYI", "ITFA PAYI"])
        isletme_nk = v(nakit,   yil, ADAY_ISLETME_NAKIT)
        yatirim_nk = v(nakit,   yil, ADAY_YATIRIM_NAKIT)

        fin_borc = _satir_deger_satirlari_toplam(bilanco, yil, ["FINANSAL BORCLAR", "FINANSAL BORÇLAR"])
        if fin_borc is None:
            if fin_kv is not None and fin_uv is not None:
                fin_borc = fin_kv + fin_uv
            elif fin_kv is not None:
                fin_borc = fin_kv
            elif fin_uv is not None:
                fin_borc = fin_uv

        net_borc = (fin_borc - nakit_) if (fin_borc is not None and nakit_ is not None) else None
        favok    = (faaliyet_k + amort) if (faaliyet_k is not None and amort is not None) else faaliyet_k
        cari_oran     = (donen / kv_yuk)      if (donen and kv_yuk)  else None
        yuk_toplam_v = (kv_yuk + uv_yuk) if (kv_yuk is not None and uv_yuk is not None) else None
        if yuk_toplam_v is None and bilanco is not None:
            yuk_toplam_v = v(bilanco, yil, ADAY_TOPLAM_YUK)
        borc_ozk      = (yuk_toplam_v / ozkay) if (yuk_toplam_v is not None and ozkay not in (None, 0)) else None
        net_borc_favok = (net_borc / favok)   if (net_borc is not None and favok and favok != 0) else None
        brut_marj     = (brut_kar / hasilat * 100) if (brut_kar and hasilat) else None
        faaliyet_marj = (faaliyet_k / hasilat * 100) if (faaliyet_k and hasilat) else None
        net_marj      = (net_kar / hasilat * 100)    if (net_kar and hasilat)    else None
        roa           = (net_kar / toplam_var * 100) if (net_kar and toplam_var) else None
        roe           = (net_kar / ozkay * 100)      if (net_kar and ozkay)      else None

        return dict(
            donen=donen, kv_yuk=kv_yuk, uv_yuk=uv_yuk, toplam_var=toplam_var,
            ozkay=ozkay, nakit_=nakit_, stok=stok, fin_borc=fin_borc,
            hasilat=hasilat, brut_kar=brut_kar, faaliyet_k=faaliyet_k,
            net_kar=net_kar, amort=amort, isletme_nk=isletme_nk,
            yatirim_nk=yatirim_nk, net_borc=net_borc, favok=favok,
            cari_oran=cari_oran, borc_ozk=borc_ozk,
            net_borc_favok=net_borc_favok, brut_marj=brut_marj,
            faaliyet_marj=faaliyet_marj, net_marj=net_marj, roa=roa, roe=roe,
        )

    e = veri_seti(yil_eski)
    y = veri_seti(yil_yeni)

    # ── Yorum üretici yardımcılar ─────────────────────────────────────────────
    def trend(eski_v, yeni_v, pozitif_iyi=True):
        d = degisim(eski_v, yeni_v)
        if d is None: return "değişim hesaplanamadı"
        if abs(d) < 2: return "yatay seyretti"
        yon = "arttı" if d > 0 else "azaldı"
        guc = "güçlü biçimde " if abs(d) > 20 else ("belirgin şekilde " if abs(d) > 8 else "")
        iyi = (d > 0) == pozitif_iyi
        renk = "olumlu" if iyi else "olumsuz"
        return f"{guc}{yon} ({fmt_pct(d)}) — {renk} bir gelişme"

    def cari_yorum(oran):
        if oran is None: return "hesaplanamadı"
        if oran >= 2.0:   return f"{fmt_oran(oran)} — güçlü likidite, kısa vadeli yükümlülükler rahatça karşılanabilir"
        if oran >= 1.5:   return f"{fmt_oran(oran)} — yeterli likidite"
        if oran >= 1.0:   return f"{fmt_oran(oran)} — sınırda likidite, yakın takip gerektirir"
        return f"{fmt_oran(oran)} — likidite baskısı var, kısa vadeli yükümlülükler dönen varlıkları aşıyor"

    def borc_ozk_yorum(oran):
        if oran is None: return "hesaplanamadı"
        if oran <= 0.5:  return f"{fmt_oran(oran)} — düşük kaldıraç, finansal risk az"
        if oran <= 1.0:  return f"{fmt_oran(oran)} — makul kaldıraç seviyesi"
        if oran <= 2.0:  return f"{fmt_oran(oran)} — yüksek kaldıraç, borç yönetimi kritik"
        return f"{fmt_oran(oran)} — çok yüksek kaldıraç, finansal kırılganlık riski mevcut"

    def net_borc_favok_yorum(oran):
        if oran is None: return "hesaplanamadı"
        if oran <= 0:    return f"{fmt_oran(oran)} — net nakit pozisyonunda, borçsuz"
        if oran <= 1.5:  return f"{fmt_oran(oran)} — düşük borç yükü, FAVÖK ile hızlı geri ödeme kapasitesi"
        if oran <= 3.0:  return f"{fmt_oran(oran)} — yönetilebilir borç yükü"
        if oran <= 5.0:  return f"{fmt_oran(oran)} — yüksek borç yükü, nakit akışı baskı altında"
        return f"{fmt_oran(oran)} — kritik borç yükü, refinansman riski"

    def marj_yorum(marj, tur):
        if marj is None: return "hesaplanamadı"
        esikler = {
            "brüt":     [(40,"güçlü"),   (25,"orta"),  (0,"zayıf")],
            "faaliyet": [(20,"güçlü"),   (10,"orta"),  (0,"zayıf")],
            "net":      [(15,"güçlü"),   (5,"orta"),   (0,"zayıf")],
        }
        for esik, etiket in esikler.get(tur, []):
            if marj >= esik:
                return f"%{marj:.1f} — {etiket} {tur} kâr marjı"
        return f"%{marj:.1f} — negatif {tur} kâr marjı, zarar söz konusu"

    # ── Bölüm bölüm analiz metni üret ────────────────────────────────────────
    satirlar = []

    def ekle(metin, bas=False):
        satirlar.append({"Kalem": metin, yil_eski: None, yil_yeni: None, "Değişim %": None})

    def baslik(metin):
        satirlar.append({"Kalem": f"▶  {metin}", yil_eski: None, yil_yeni: None, "Değişim %": None})

    def bos():
        satirlar.append({"Kalem": "", yil_eski: None, yil_yeni: None, "Değişim %": None})

    # ── 1. GENEL DURUM ────────────────────────────────────────────────────────
    baslik(f"1. GENEL DURUM  ({yil_eski} → {yil_yeni})")
    var_trend  = trend(e["toplam_var"], y["toplam_var"])
    has_trend  = trend(e["hasilat"],    y["hasilat"])
    kar_trend  = trend(e["net_kar"],    y["net_kar"])
    ekle(f"  • Toplam varlıklar {var_trend}. Dönem sonu: {fmt_para(y['toplam_var'])}.")
    ekle(f"  • Net hasılat {has_trend}. Dönem sonu: {fmt_para(y['hasilat'])}.")
    if y["net_kar"] is not None:
        ekle(f"  • Net kâr/zarar {kar_trend}. Dönem sonu: {fmt_para(y['net_kar'])}.")
    bos()

    # ── 2. LİKİDİTE ──────────────────────────────────────────────────────────
    baslik("2. LİKİDİTE ANALİZİ")
    ekle(f"  • {yil_eski}: Cari oran {cari_yorum(e['cari_oran'])}")
    ekle(f"  • {yil_yeni}: Cari oran {cari_yorum(y['cari_oran'])}")
    if e["nakit_"] and y["nakit_"]:
        ekle(f"  • Nakit pozisyonu {trend(e['nakit_'], y['nakit_'])}. Dönem sonu: {fmt_para(y['nakit_'])}.")
    bos()

    # ── 3. BORÇLULUK ─────────────────────────────────────────────────────────
    baslik("3. BORÇLULUK & FİNANSAL KALDIRAÇ")
    ekle(f"  • {yil_eski}: Borç/Özkaynak {borc_ozk_yorum(e['borc_ozk'])}")
    ekle(f"  • {yil_yeni}: Borç/Özkaynak {borc_ozk_yorum(y['borc_ozk'])}")
    ekle(f"  • {yil_eski}: Net Borç/FAVÖK {net_borc_favok_yorum(e['net_borc_favok'])}")
    ekle(f"  • {yil_yeni}: Net Borç/FAVÖK {net_borc_favok_yorum(y['net_borc_favok'])}")
    if y["fin_borc"] is not None:
        ekle(f"  • Toplam finansal borç: {fmt_para(y['fin_borc'])}.")
    bos()

    # ── 4. BÜYÜME ────────────────────────────────────────────────────────────
    baslik("4. BÜYÜME ANALİZİ")
    ekle(f"  • Hasılat büyümesi: {trend(e['hasilat'], y['hasilat'])}")
    ekle(f"  • Varlık büyümesi: {trend(e['toplam_var'], y['toplam_var'])}")
    ekle(f"  • Özkaynak büyümesi: {trend(e['ozkay'], y['ozkay'])}")
    if e["favok"] and y["favok"]:
        ekle(f"  • FAVÖK büyümesi: {trend(e['favok'], y['favok'])}. Dönem sonu: {fmt_para(y['favok'])}.")
    bos()

    # ── 5. KARLILIK ───────────────────────────────────────────────────────────
    baslik("5. KARLILIK ANALİZİ")
    ekle(f"  • {yil_yeni} brüt kâr marjı: {marj_yorum(y['brut_marj'], 'brüt')}")
    ekle(f"  • {yil_yeni} faaliyet kâr marjı: {marj_yorum(y['faaliyet_marj'], 'faaliyet')}")
    ekle(f"  • {yil_yeni} net kâr marjı: {marj_yorum(y['net_marj'], 'net')}")
    if y["roa"] is not None:
        ekle(f"  • Aktif kârlılığı (ROA): %{y['roa']:.2f} — her 100₺ varlıktan {y['roa']:.2f}₺ net kâr üretildi.")
    if y["roe"] is not None:
        ekle(f"  • Özkaynak kârlılığı (ROE): %{y['roe']:.2f}")
    bos()

    # ── 6. NAKİT AKIŞI ───────────────────────────────────────────────────────
    baslik("6. NAKİT AKIŞI")
    if y["isletme_nk"] is not None:
        poz = "pozitif — şirket kendi kendini fonlayabiliyor" if y["isletme_nk"] > 0 else "negatif — dış finansman ihtiyacı var"
        ekle(f"  • İşletme faaliyetlerinden nakit akışı: {fmt_para(y['isletme_nk'])} ({poz})")
    if y["yatirim_nk"] is not None:
        poz = "negatif — yatırım yapılıyor (büyüme sinyali)" if y["yatirim_nk"] < 0 else "pozitif — varlık satışı söz konusu"
        ekle(f"  • Yatırım faaliyetlerinden nakit akışı: {fmt_para(y['yatirim_nk'])} ({poz})")
    bos()

    # ── 7. GÜÇLÜ YÖNLER ──────────────────────────────────────────────────────
    baslik("7. GÜÇLÜ YÖNLER")
    guclu = []
    if y["cari_oran"] and y["cari_oran"] >= 1.5:
        guclu.append("  ✔  Güçlü likidite pozisyonu — kısa vadeli yükümlülükler rahatça karşılanabilir durumda")
    if y["net_borc_favok"] is not None and y["net_borc_favok"] <= 2.0:
        guclu.append("  ✔  Düşük borç yükü — FAVÖK ile borçları hızla geri ödeme kapasitesi mevcut")
    if y["net_marj"] and y["net_marj"] > 10:
        guclu.append(f"  ✔  Güçlü net kâr marjı (%{y['net_marj']:.1f}) — sektör ortalamasının üzerinde kârlılık")
    if y["favok"] and e["favok"] and y["favok"] > e["favok"]:
        guclu.append(f"  ✔  FAVÖK artışı — nakit yaratma kapasitesi dönem dönem güçleniyor")
    if y["hasilat"] and e["hasilat"] and y["hasilat"] > e["hasilat"] * 1.1:
        guclu.append(f"  ✔  Güçlü hasılat büyümesi ({fmt_pct(degisim(e['hasilat'], y['hasilat']))}) — pazar payı korunuyor")
    if y["nakit_"] and y["nakit_"] > 0:
        guclu.append(f"  ✔  Nakit ve nakit benzerleri: {fmt_para(y['nakit_'])} — güçlü nakit tamponu")
    if not guclu:
        guclu.append("  • Mevcut verilerle belirgin güçlü yön tespit edilemedi")
    for g in guclu[:4]: ekle(g)
    bos()

    # ── 8. RİSKLER ───────────────────────────────────────────────────────────
    baslik("8. RİSKLER & DİKKAT NOKTALARI")
    riskler = []
    if y["cari_oran"] and y["cari_oran"] < 1.0:
        riskler.append("  ⚠  Likidite riski — kısa vadeli yükümlülükler dönen varlıkları aşıyor")
    if y["borc_ozk"] and y["borc_ozk"] > 2.0:
        riskler.append(f"  ⚠  Yüksek kaldıraç riski — Borç/Özkaynak oranı {fmt_oran(y['borc_ozk'])}")
    if y["net_borc_favok"] and y["net_borc_favok"] > 4.0:
        riskler.append(f"  ⚠  Borç ödeme kapasitesi baskı altında — Net Borç/FAVÖK: {fmt_oran(y['net_borc_favok'])}")
    if y["net_kar"] is not None and y["net_kar"] < 0:
        riskler.append("  ⚠  Net zarar — şirket dönemde kâr üretemedi")
    if y["isletme_nk"] is not None and y["isletme_nk"] < 0:
        riskler.append("  ⚠  Negatif işletme nakit akışı — dış finansman bağımlılığı artıyor")
    if y["hasilat"] and e["hasilat"] and y["hasilat"] < e["hasilat"]:
        riskler.append(f"  ⚠  Hasılat daralması ({fmt_pct(degisim(e['hasilat'], y['hasilat']))}) — gelir kaybı yaşanıyor")
    if y["brut_marj"] and e["brut_marj"] and y["brut_marj"] < e["brut_marj"] - 3:
        riskler.append("  ⚠  Brüt kâr marjı geriledi — maliyet baskısı veya fiyatlama sorunu")
    if not riskler:
        riskler.append("  • Temel finansal göstergeler itibarıyla belirgin risk tespit edilmedi")
    for r in riskler[:4]: ekle(r)
    bos()

    # ── 9. GENEL DEĞERLENDİRME ───────────────────────────────────────────────
    baslik("9. GENEL DEĞERLENDİRME")
    puan = 0
    if y["cari_oran"] and y["cari_oran"] >= 1.5: puan += 1
    if y["borc_ozk"] and y["borc_ozk"] <= 1.5:  puan += 1
    if y["net_kar"] and y["net_kar"] > 0:        puan += 1
    if y["isletme_nk"] and y["isletme_nk"] > 0: puan += 1
    if y["hasilat"] and e["hasilat"] and y["hasilat"] > e["hasilat"]: puan += 1
    if y["net_borc_favok"] is not None and y["net_borc_favok"] <= 2.5: puan += 1

    if puan >= 5:
        genel = "güçlü finansal yapıya sahip"
        tavsiye = "Finansal göstergeler genel itibarıyla olumlu seyretmekte."
    elif puan >= 3:
        genel = "orta düzey finansal sağlığa sahip"
        tavsiye = "Bazı alanlarda iyileşme potansiyeli bulunmakta, gelişmeler yakından takip edilmeli."
    else:
        genel = "finansal baskı altında"
        tavsiye = "Likidite, borçluluk veya kârlılık alanlarında dikkat gerektiren göstergeler mevcut."

    ekle(f"  Şirket {yil_yeni} itibarıyla {genel} görünmektedir. {tavsiye}")
    ekle(f"  Değerlendirme skoru: {puan}/6 temel gösterge olumlu.")
    bos()
    ekle("  ⚠  Bu analiz otomatik hesaplama motoruyla üretilmiştir. Yatırım kararlarında")
    ekle("     uzman görüşü ve detaylı araştırma yapılması tavsiye edilir.")

    log(f"  ✅ Yerleşik AI analizi tamamlandı ({len(satirlar)} satır)")
    return satirlar

def sirala(df, donem_sutunlar):
    if len(donem_sutunlar) < 2:
        e = donem_sutunlar[0] if donem_sutunlar else ""
        df = df.copy(); df["Değişim %"] = None; return df, e, e

    def yil_anahtari(s):
        m = re.search(r'(\d{4})', str(s))
        return int(m.group(1)) if m else 0

    sirali = sorted(donem_sutunlar, key=yil_anahtari, reverse=True)  # en yeni solda
    eski = sirali[-1]   # en eski
    yeni = sirali[0]    # en yeni

    # Değişim sütunları ekle
    degisim_sutunlari = []
    for i in range(len(sirali)-1):
        yeni_yil = sirali[i]
        eski_yil = sirali[i+1]
        sutun_adi = f"{yeni_yil}-{eski_yil} %"
        df[sutun_adi] = None
        degisim_sutunlari.append(sutun_adi)

    diger = [c for c in df.columns if c not in ("Kalem",) + tuple(sirali) + tuple(degisim_sutunlari)]
    df = df.reindex(columns=["Kalem"] + sirali + degisim_sutunlari + diger)

    # Değişim % Python'da hesapla — Excel formülü sekme_yaz'da yazılacak (eski kod için)
    # Artık birden fazla değişim sütunu var, ama eski için bırakalım
    df["Değişim %"] = None  # sekme_yaz'da Excel formülü yazılacak (eğer varsa)
    return df, eski, yeni

def sirala_analiz(df, donem_sutunlar):
    """
    Analiz sekmesi için: satır sırasını korur (oranlar + AI metni karışmaz).
    Değişim % değerleri sekme_yaz'da Excel formülüyle yazılır; burada sütun düzeni sağlanır.
    """
    if len(donem_sutunlar) < 2:
        e = donem_sutunlar[0] if donem_sutunlar else ""
        df = df.copy()
        if "Değişim %" not in df.columns:
            df["Değişim %"] = None
        return df, e, e

    def yil_anahtari(s):
        m = re.search(r'(\d{4})', str(s))
        return int(m.group(1)) if m else 0

    sirali = sorted(donem_sutunlar, key=yil_anahtari, reverse=True)  # en yeni solda
    eski = sirali[-1]
    yeni = sirali[0]

    # Değişim sütunları ekle
    degisim_sutunlari = []
    for i in range(len(sirali)-1):
        yeni_yil = sirali[i]
        eski_yil = sirali[i+1]
        sutun_adi = f"{yeni_yil}-{eski_yil} %"
        df[sutun_adi] = None
        degisim_sutunlari.append(sutun_adi)

    diger = [c for c in df.columns if c not in ("Kalem",) + tuple(sirali) + tuple(degisim_sutunlari)]
    df = df.reindex(columns=["Kalem"] + sirali + degisim_sutunlari + diger)
    df = df.copy()
    df["Değişim %"] = None  # eski için
    return df, eski, yeni

def sekme_yaz(wb, df, sekme_adi, hisse, eski, yeni):
    renk = SEKME_RENK[sekme_adi]; ana_kw = ANA_KW.get(sekme_adi,[])
    ws = wb.create_sheet(sekme_adi)
    brd = Border(left=Side(style="thin",color="CCCCCC"),right=Side(style="thin",color="CCCCCC"),
                 top=Side(style="thin",color="CCCCCC"),bottom=Side(style="thin",color="CCCCCC"))

    ws.merge_cells(f"A1:{get_column_letter(len(df.columns))}1")
    ws["A1"] = f"{hisse} – {sekme_adi} | {eski} → {yeni} | Kaynak: İş Yatırım"
    ws["A1"].font=Font(name="Arial",bold=True,size=13,color="FFFFFF")
    ws["A1"].fill=PatternFill("solid",fgColor=renk)
    ws["A1"].alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height=30

    for ci,cn in enumerate(df.columns,1):
        h=ws.cell(row=2,column=ci,value=str(cn))
        h.font=Font(name="Arial",bold=True,size=10,color="FFFFFF")
        h.fill=PatternFill("solid",fgColor=renk)
        h.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        h.border=brd
    ws.row_dimensions[2].height=22

    # Değişim % için B ve C sütun harflerini bul (eski=B, yeni=C)
    cols_list = list(df.columns)
    eski_ci = cols_list.index(eski) + 1 if eski in cols_list else None
    yeni_ci = cols_list.index(yeni) + 1 if yeni in cols_list else None
    degisim_ci = cols_list.index("Değişim %") + 1 if "Değişim %" in cols_list else None
    eski_harf = get_column_letter(eski_ci) if eski_ci else "B"
    yeni_harf = get_column_letter(yeni_ci) if yeni_ci else "C"

    for ri,(_,satir) in enumerate(df.iterrows(),3):
        kv_meta = satir["Kalem"]
        if kv_meta is None or (isinstance(kv_meta, float) and pd.isna(kv_meta)):
            kalem_str = ""
        else:
            kalem_str = str(kv_meta)
        kalem = kalem_str.upper() if sekme_adi != "Analiz" else kalem_str
        ana=any(k in kalem.upper() for k in ana_kw)

        if sekme_adi == "Analiz":
            is_baslik = (kalem_str.strip().startswith(("1.","2.","3.","4.","5.","6.","7.","8.","9.","##","▶")) or
                         (len(kalem_str.strip()) > 0 and kalem_str.strip().isupper()))
            dolgu = "D6E4F0" if is_baslik else ("F2F2F2" if ri%2==0 else "FFFFFF")
        else:
            dolgu="D6E4F0" if ana else ("F2F2F2" if ri%2==0 else "FFFFFF")

        # Analiz: sayı yok, sadece metin → tüm sütunları birleştir (boş B/C/D görünümü düzelir)
        if sekme_adi == "Analiz":
            bn = sayiya_cevir_float(satir.get(eski))
            cn = sayiya_cevir_float(satir.get(yeni))
            if bn is None and cn is None and kalem_str.strip() != "":
                last = get_column_letter(len(df.columns))
                ws.merge_cells(f"A{ri}:{last}{ri}")
                c = ws.cell(row=ri, column=1, value=kalem_str)
                c.font = Font(name="Arial", bold=(ana or is_baslik), size=9)
                c.fill = PatternFill("solid", fgColor=dolgu)
                c.border = brd
                c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                if is_baslik:
                    ws.row_dimensions[ri].height = 22
                continue

        for ci,(col,val) in enumerate(satir.items(),1):
            if col=="Kalem":
                kv = satir["Kalem"]
                if kv is None or (isinstance(kv, float) and pd.isna(kv)):
                    gosterim = ""
                else:
                    gosterim = str(kv)
                fmt = None
                hiz = "left"
            elif col.endswith(" %") and "-" in col:
                # Yıldan yıla değişim formülü
                parts = col.replace(" %", "").split("-")
                if len(parts) == 2:
                    yeni_yil, eski_yil = parts
                    yeni_ci_col = cols_list.index(yeni_yil) + 1 if yeni_yil in cols_list else None
                    eski_ci_col = cols_list.index(eski_yil) + 1 if eski_yil in cols_list else None
                    if yeni_ci_col and eski_ci_col:
                        yeni_harf_col = get_column_letter(yeni_ci_col)
                        eski_harf_col = get_column_letter(eski_ci_col)
                        gosterim = f"=({yeni_harf_col}{ri}-{eski_harf_col}{ri})/ABS({eski_harf_col}{ri})"
                        fmt = '0.00%'
                    else:
                        gosterim = None
                        fmt = '0.00%'
                else:
                    gosterim = None
                    fmt = '0.00%'
                hiz = "center"
            elif col=="Değişim %":
                # Excel formülü: =(C3-B3)/ABS(B3) — oran satırlarında; Analiz metin satırlarında boş
                if sekme_adi == "Analiz" and eski_ci and yeni_ci:
                    b_val = sayiya_cevir_float(satir.get(eski))
                    c_val = sayiya_cevir_float(satir.get(yeni))
                    if b_val is not None and c_val is not None and b_val != 0:
                        gosterim = f"=({yeni_harf}{ri}-{eski_harf}{ri})/ABS({eski_harf}{ri})"
                        fmt = '0.00%'
                    else:
                        gosterim = None
                        fmt = None
                elif sekme_adi != "Analiz" and eski_ci and yeni_ci:
                    gosterim = f"=({yeni_harf}{ri}-{eski_harf}{ri})/ABS({eski_harf}{ri})"
                    fmt = '0.00%'
                else:
                    gosterim = None
                    fmt = '0.00%'
                hiz = "center"
            else:
                if sekme_adi == "Analiz":
                    gosterim = sayiya_cevir_float(val)
                    fmt, hiz = '#,##0.00_);(#,##0.00);"-"', "right"
                else:
                    gosterim = sayiya_cevir_int(val)
                    fmt, hiz = '#,##0_);(#,##0);"-"', "right"
            c=ws.cell(row=ri,column=ci,value=gosterim)
            c.font=Font(name="Arial",bold=(ana or (sekme_adi=="Analiz" and is_baslik)),size=9)
            c.fill=PatternFill("solid",fgColor=dolgu)
            c.border=brd
            c.alignment=Alignment(horizontal=hiz,vertical="center",wrap_text=(sekme_adi=="Analiz"))
            if fmt: c.number_format=fmt

    ws.column_dimensions["A"].width = 80 if sekme_adi == "Analiz" else 55
    for ci in range(2,len(df.columns)+1):
        ws.column_dimensions[get_column_letter(ci)].width=(14 if df.columns[ci-1]=="Değişim %" else 22)
    ws.freeze_panes="B3"
    ws.cell(row=ws.max_row+2,column=1,
            value="Kaynak: İş Yatırım (isyatirim.com.tr) – TFRS Verileri"
    ).font=Font(name="Arial",size=8,italic=True,color="888888")

def dosya_adi_bul(tercih):
    if not os.path.exists(tercih): return tercih
    try:
        with open(tercih,"a"): return tercih
    except PermissionError:
        ts=datetime.now().strftime("%H%M%S"); kok,uzanti=os.path.splitext(tercih)
        return f"{kok}_{ts}{uzanti}"

def _excel_sutun_ad_normalize(c):
    s = str(c).strip()
    if len(s) > 2 and s.endswith(".0"):
        alt = s[:-2]
        if alt.isdigit() or re.match(r"^\d{4}$", alt):
            return alt
    return s

def excel_den_tablolar_oku(path, log):
    """
    Bu uygulamanın ürettiği (veya aynı yapıdaki) finansal tablolar .xlsx dosyasından
    Bilanço / Gelir / Nakit / Dipnot sayfalarını okur — AI analiz motoru için.
    """
    if not os.path.isfile(path):
        raise FileNotFoundError(path)
    xl = pd.ExcelFile(path)
    tablolar = {}
    for sekme in ["Bilanço", "Gelir Tablosu", "Nakit Akım", "Dipnot"]:
        if sekme not in xl.sheet_names:
            log(f"  Sekme yok: {sekme}")
            continue
        df = pd.read_excel(path, sheet_name=sekme, header=1)
        df.columns = [_excel_sutun_ad_normalize(x) for x in df.columns]
        if "Kalem" not in df.columns and len(df.columns):
            df = df.rename(columns={df.columns[0]: "Kalem"})
        df = df.dropna(how="all")
        if "Kalem" not in df.columns:
            log(f"  ⚠️  {sekme}: 'Kalem' sütunu yok"); continue
        df["Kalem"] = df["Kalem"].astype(str).str.strip()
        df = df[df["Kalem"].str.len() > 0]
        df = df[~df["Kalem"].str.contains("Kaynak:", case=False, na=False)]
        donem = [c for c in df.columns if c not in ("Kalem", "Değişim %")]
        tablolar[sekme] = (df.reset_index(drop=True), donem)
    return tablolar

def analiz_donemleri_bul(tablolar):
    bilanco = tablolar.get("Bilanço", (None, None))[0]
    if bilanco is None or bilanco.empty:
        return None, None
    gercek = sorted(
        [c for c in bilanco.columns if c != "Kalem" and re.search(r"\d{4}", str(c))],
        key=lambda s: int(re.search(r"\d{4}", str(s)).group()) if re.search(r"\d{4}", str(s)) else 0
    )
    if len(gercek) < 1:
        return None, None
    return gercek[0], gercek[-1]

def hisse_kodu_excel_yolundan(path):
    base = os.path.basename(path or "")
    m = re.match(r"^([A-Z0-9]+)_finansal", base.upper())
    if m:
        return m.group(1)
    stem = os.path.splitext(base)[0]
    return stem.split("_")[0].upper() if stem else "?"

def llm_ek_ozet_tr(hisse, analiz_metni, log):
    """OPENAI_API_KEY ortam değişkeni varsa kısa Türkçe yönetici özeti üretir."""
    key = os.environ.get("OPENAI_API_KEY", "").strip()
    if not key:
        return None
    try:
        log("  LLM (OpenAI) kısa özet isteniyor...")
        payload = {
            "model": "gpt-4o-mini",
            "messages": [
                {"role": "system", "content": "Sen deneyimli bir finansal analistsin. Türkçe, kısa ve net yaz; yatırım tavsiyesi verme, sadece veriye dayalı özet."},
                {"role": "user", "content": f"Hisse: {hisse}\n\nAşağıdaki otomatik finansal analize dayanarak 5–8 cümlelik yönetici özeti yaz:\n\n{analiz_metni[:14000]}"}
            ],
            "max_tokens": 1200,
            "temperature": 0.25,
        }
        r = req_lib.post(
            "https://api.openai.com/v1/chat/completions",
            headers={"Authorization": f"Bearer {key}", "Content-Type": "application/json"},
            json=payload,
            timeout=120,
        )
        r.raise_for_status()
        data = r.json()
        return data["choices"][0]["message"]["content"].strip()
    except Exception as e:
        log(f"  ⚠️  LLM özet atlandı: {e}")
        return None

def analiz_metni_olustur(tablolar, yil_eski, yil_yeni, hisse, log, llm_istek=True):
    """Oran tablosu + yerleşik AI metni; isteğe bağlı OpenAI özeti."""
    analiz_oran = analiz_df_olustur(tablolar, yil_eski, yil_yeni)
    ai_satirlar = yerlesik_ai_analiz(tablolar, yil_eski, yil_yeni, log)
    lines = []
    lines.append("═" * 56)
    lines.append(f"  FİNANSAL ORANLAR  ({yil_eski}  →  {yil_yeni})  |  {hisse}")
    lines.append("═" * 56)
    if analiz_oran is not None and not analiz_oran.empty:
        for _, row in analiz_oran.iterrows():
            k = row.get("Kalem", "")
            a, b = row.get(yil_eski), row.get(yil_yeni)
            d = row.get("Değişim %")
            lines.append(f"  • {k}")
            lines.append(f"      {yil_eski}: {a}  |  {yil_yeni}: {b}  |  Değişim %: {d}")
    else:
        lines.append("  (Oran tablosu oluşturulamadı — Bilanço / Gelir / Nakit verisi eksik olabilir.)")
    lines.append("")
    lines.append("═" * 56)
    lines.append("  YAPAY ZEKA FİNANSAL ANALİZİ  (yerleşik kural motoru)")
    lines.append("═" * 56)
    lines.append("")
    if ai_satirlar:
        for s in ai_satirlar:
            lines.append(str(s.get("Kalem", "")))
    metin_govde = "\n".join(lines)
    if llm_istek:
        oz = llm_ek_ozet_tr(hisse, metin_govde, log)
        if oz:
            lines.append("")
            lines.append("═" * 56)
            lines.append("  LLM YÖNETİCİ ÖZETİ  (OpenAI — OPENAI_API_KEY)")
            lines.append("═" * 56)
            lines.append(oz)
    return "\n".join(lines)

# ─── Ana iş mantığı (thread'den çağrılır) ─────────────────────────────────────
def isle(hisse, kayit_klasoru, log, bitti_cb, hata_cb, baslangic=BASLANGIC, bitis=BITIS, donem_ay=DONEM_AY):
    try:
        log(f"{'='*50}")
        log(f"  {hisse}  |  {baslangic} → {bitis}  |  Dönem: {donem_ay}")
        log(f"{'='*50}")

        log("\n[1] Hisse türü belirleniyor...")
        tur = hisse_turu_belirle(hisse, log)

        wb = Workbook(); wb.remove(wb.active)
        global_eski = str(baslangic); global_yeni = str(bitis)

        log("\n[2] Veriler çekiliyor...")
        if tur == "sanayi":
            tablolar = cek_sanayi(hisse, baslangic, bitis, log)
            if str(donem_ay) not in {"12", "ALL"}:
                log("  ℹ️  Sanayi (borsapy) kaynağı çeyrek dönem seçimini desteklemiyor; yıllık sütunlar kullanıldı.")
        else:
            tablolar = cek_finans(hisse, baslangic, bitis, log, donem_ay=donem_ay)

        # Dipnot: her iki hisse türü için dipnot_cek ile çek
        log("\n[3] Dipnot çekiliyor...")
        df_dipnot, donem_d = dipnot_cek(hisse, baslangic, bitis, log, donem_ay=donem_ay)
        if df_dipnot is not None and not df_dipnot.empty:
            tablolar["Dipnot"] = (df_dipnot, donem_d)

        log("\n[4] Excel oluşturuluyor...")
        # Analiz için gerçek sütun adlarını bul (rename sonrası ne olduğunu al)
        bilanco_df_kontrol = tablolar.get("Bilanço", (None, None))[0]
        gercek_donemler = []
        if bilanco_df_kontrol is not None:
            gercek_donemler = sorted(
                [c for c in bilanco_df_kontrol.columns if c != "Kalem"],
                key=lambda s: int(re.search(r'\d{4}', str(s)).group()) if re.search(r'\d{4}', str(s)) else 0
            )

        analiz_yil_eski = gercek_donemler[0]  if len(gercek_donemler) >= 1 else str(baslangic)
        analiz_yil_yeni = gercek_donemler[-1] if len(gercek_donemler) >= 1 else str(bitis)
        log(f"  Analiz dönemleri: {analiz_yil_eski} → {analiz_yil_yeni}")

        # Analiz sekmesi: finansal oranlar + Claude API yorumu
        try:
            analiz_oran = analiz_df_olustur(tablolar, analiz_yil_eski, analiz_yil_yeni)
            ai_satirlar = yerlesik_ai_analiz(tablolar, analiz_yil_eski, analiz_yil_yeni, log)

            satirlar_birlesik = []
            if analiz_oran is not None and not analiz_oran.empty:
                for _, row in analiz_oran.iterrows():
                    satirlar_birlesik.append(row.to_dict())

            if ai_satirlar:
                satirlar_birlesik.append({"Kalem": "", analiz_yil_eski: None, analiz_yil_yeni: None, "Değişim %": None})
                satirlar_birlesik.append({"Kalem": "━━━ YAPAY ZEKA FİNANSAL ANALİZİ ━━━", analiz_yil_eski: None, analiz_yil_yeni: None, "Değişim %": None})
                satirlar_birlesik.append({"Kalem": "", analiz_yil_eski: None, analiz_yil_yeni: None, "Değişim %": None})
                satirlar_birlesik.extend(ai_satirlar)

            if satirlar_birlesik:
                analiz_df = pd.DataFrame(satirlar_birlesik)
                for col in [analiz_yil_eski, analiz_yil_yeni, "Değişim %"]:
                    if col not in analiz_df.columns:
                        analiz_df[col] = None
                analiz_df = analiz_df[["Kalem", analiz_yil_eski, analiz_yil_yeni, "Değişim %"]]
                tablolar["Analiz"] = (analiz_df, [analiz_yil_eski, analiz_yil_yeni])
            elif analiz_oran is not None and not analiz_oran.empty:
                tablolar["Analiz"] = (analiz_oran, [analiz_yil_eski, analiz_yil_yeni])
        except Exception as e:
            log(f"  ⚠️  Analiz hesaplanamadı: {e}")

        for sekme in ["Analiz","Bilanço","Gelir Tablosu","Nakit Akım","Dipnot"]:
            if sekme not in tablolar: log(f"  ⏭  '{sekme}' — veri yok"); continue
            df,donem = tablolar[sekme]
            if sekme == "Analiz":
                df_son, eski, yeni = sirala_analiz(df, donem)
            else:
                df_son, eski, yeni = sirala(df, donem)
            global_eski,global_yeni = eski,yeni
            sekme_yaz(wb,df_son,sekme,hisse,eski,yeni)
            log(f"  📋 '{sekme}' → {len(df_son)} kalem")

        if not wb.sheetnames:
            hata_cb("Hiç veri alınamadı.\nHisse kodunu kontrol edin."); return

        cikti = os.path.join(kayit_klasoru, f"{hisse}_finansal_tablolar.xlsx")
        cikti = dosya_adi_bul(cikti)
        wb.save(cikti)

        log(f"\n{'='*50}")
        log(f"✅ Tamamlandı!")
        log(f"📁 {cikti}")
        bitti_cb(cikti)

    except Exception as e:
        hata_cb(f"{e}\n\n{traceback.format_exc()}")

# ══════════════════════════════════════════════════════════════════════════════
# 5.  GUI
# ══════════════════════════════════════════════════════════════════════════════

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("BIST Finansal Tablolar")
        self.geometry("720x620")
        self.minsize(640, 520)
        self.resizable(True, True)
        self.configure(bg="#F0F4F8")
        self._son_dosya = None
        self._yukle_arayuz()

    def _yukle_arayuz(self):
        baslik = tk.Frame(self, bg="#1F4E79", padx=20, pady=14)
        baslik.pack(fill="x")
        tk.Label(baslik, text="📊  BIST Finansal Tablolar",
                 font=("Arial", 16, "bold"), fg="white", bg="#1F4E79").pack(side="left")
        self.donem_ust_etiket = tk.Label(
            baslik, text=f"{BASLANGIC} → {BITIS} | 12",
            font=("Arial", 11), fg="#A8C8E8", bg="#1F4E79"
        )
        self.donem_ust_etiket.pack(side="right")

        self._notebook = ttk.Notebook(self)
        self._notebook.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        tab_veri = tk.Frame(self._notebook, bg="#F0F4F8")
        tab_analiz = tk.Frame(self._notebook, bg="#F0F4F8")
        self._notebook.add(tab_veri, text="  Veri çek & Excel  ")
        self._notebook.add(tab_analiz, text="  AI Analiz sayfası  ")

        # ── Sekme 1: Veri çek ─────────────────────────────────────────────
        giris = tk.Frame(tab_veri, bg="#F0F4F8", pady=18)
        giris.pack(fill="x", padx=24)

        tk.Label(giris, text="Hisse Kodu:", font=("Arial",11,"bold"),
                 bg="#F0F4F8", fg="#333").grid(row=0, column=0, sticky="w")
        self.hisse_var = tk.StringVar()
        self.hisse_giris = tk.Entry(giris, textvariable=self.hisse_var,
                                    font=("Arial",13), width=12,
                                    bd=2, relief="groove")
        self.hisse_giris.grid(row=0, column=1, padx=(10,6))
        self.hisse_giris.bind("<Return>", lambda e: self._baslat())

        tk.Label(giris, text="örn: THYAO, AKBNK, EKGYO",
                 font=("Arial",9), fg="#888", bg="#F0F4F8").grid(row=0, column=2, sticky="w")

        donem_frame = tk.Frame(tab_veri, bg="#F0F4F8")
        donem_frame.pack(fill="x", padx=24, pady=(0,8))
        tk.Label(donem_frame, text="Yıl Aralığı:", font=("Arial",10),
                 bg="#F0F4F8", fg="#555").pack(side="left")
        mevcut_yil = datetime.now().year
        yil_list = [str(y) for y in range(mevcut_yil - 8, mevcut_yil + 1)]
        self.baslangic_var = tk.StringVar(value=str(BASLANGIC))
        self.bitis_var = tk.StringVar(value=str(BITIS))
        self.baslangic_cb = ttk.Combobox(donem_frame, textvariable=self.baslangic_var, values=yil_list, width=8, state="readonly")
        self.baslangic_cb.pack(side="left", padx=(8,4))
        tk.Label(donem_frame, text="→", font=("Arial",10), bg="#F0F4F8", fg="#666").pack(side="left", padx=2)
        self.bitis_cb = ttk.Combobox(donem_frame, textvariable=self.bitis_var, values=yil_list, width=8, state="readonly")
        self.bitis_cb.pack(side="left", padx=(4,10))

        tk.Label(donem_frame, text="Dönem:", font=("Arial",10),
                 bg="#F0F4F8", fg="#555").pack(side="left")
        self.donem_map = {lbl: val for lbl, val in DONEM_SECENEKLERI}
        self.donem_label_var = tk.StringVar(value=DONEM_SECENEKLERI[0][0])
        self.donem_cb = ttk.Combobox(
            donem_frame, textvariable=self.donem_label_var,
            values=[lbl for lbl, _ in DONEM_SECENEKLERI], width=16, state="readonly"
        )
        self.donem_cb.pack(side="left", padx=(8,0))

        klasor_frame = tk.Frame(tab_veri, bg="#F0F4F8")
        klasor_frame.pack(fill="x", padx=24, pady=(0,10))
        tk.Label(klasor_frame, text="Kayıt klasörü:", font=("Arial",10),
                 bg="#F0F4F8", fg="#555").pack(side="left")
        self.klasor_var = tk.StringVar(value=os.path.expanduser("~\\Desktop"))
        tk.Entry(klasor_frame, textvariable=self.klasor_var,
                 font=("Arial",9), width=38, bd=1, relief="groove").pack(side="left", padx=(8,4))
        tk.Button(klasor_frame, text="📂 Seç", font=("Arial",9),
                  command=self._klasor_sec, relief="groove", cursor="hand2").pack(side="left")

        ai_frame = tk.Frame(tab_veri, bg="#E8F5E9", bd=1, relief="solid")
        ai_frame.pack(fill="x", padx=24, pady=(0,8))
        tk.Label(ai_frame,
                 text="🤖  Excel’deki Analiz sekmesi: yerleşik AI motoru + isteğe bağlı OpenAI özeti (AI Analiz sekmesi)",
                 font=("Arial", 8), fg="#2E7D32", bg="#E8F5E9", pady=4).pack(anchor="w", padx=8)

        self.btn = tk.Button(tab_veri, text="▶  Excel Oluştur",
                             font=("Arial",12,"bold"), bg="#1F4E79", fg="white",
                             activebackground="#163A5F", activeforeground="white",
                             bd=0, padx=22, pady=8, cursor="hand2",
                             command=self._baslat)
        self.btn.pack(pady=(0,6))

        self.pb = ttk.Progressbar(tab_veri, mode="indeterminate", length=500)
        self.pb.pack(pady=(0,8))

        log_frame = tk.Frame(tab_veri, bg="#F0F4F8")
        log_frame.pack(fill="both", expand=True, padx=24, pady=(0,12))
        tk.Label(log_frame, text="İşlem Günlüğü", font=("Arial",9),
                 fg="#888", bg="#F0F4F8").pack(anchor="w")
        self.log_alan = tk.Text(log_frame, font=("Consolas",9), height=12,
                                bg="#1A1A2E", fg="#E0E0E0", bd=1,
                                relief="flat", state="disabled",
                                wrap="word", padx=8, pady=6)
        sb = tk.Scrollbar(log_frame, command=self.log_alan.yview)
        self.log_alan.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self.log_alan.pack(fill="both", expand=True)

        self.sonuc_frame = tk.Frame(tab_veri, bg="#F0F4F8")
        self.sonuc_frame.pack(fill="x", padx=24, pady=(0,14))

        # ── Sekme 2: AI Analiz (Excel’deki verilerle) ─────────────────────
        self.analiz_dosya_var = tk.StringVar(value="")
        a_top = tk.Frame(tab_analiz, bg="#F0F4F8", padx=20, pady=12)
        a_top.pack(fill="x")
        tk.Label(a_top, text="Finansal tablolar Excel dosyası (.xlsx):",
                 font=("Arial", 10, "bold"), bg="#F0F4F8", fg="#333").pack(anchor="w")
        row = tk.Frame(a_top, bg="#F0F4F8")
        row.pack(fill="x", pady=(6, 0))
        tk.Entry(row, textvariable=self.analiz_dosya_var, font=("Arial", 9), width=52,
                 bd=1, relief="groove").pack(side="left", padx=(0, 6))
        tk.Button(row, text="📂 Excel seç", font=("Arial", 9), relief="groove", cursor="hand2",
                  command=self._analiz_dosya_sec).pack(side="left", padx=(0, 6))
        self.analiz_llm_var = tk.BooleanVar(value=bool(os.environ.get("OPENAI_API_KEY", "").strip()))
        tk.Checkbutton(row, text="OpenAI özet", variable=self.analiz_llm_var,
                       font=("Arial", 9), bg="#F0F4F8", fg="#333",
                       cursor="hand2").pack(side="left", padx=4)
        self.analiz_btn = tk.Button(row, text="▶  Analizi çalıştır",
                                    font=("Arial", 10, "bold"), bg="#6A1B9A", fg="white",
                                    activebackground="#5a1582", activeforeground="white",
                                    bd=0, padx=14, pady=6, cursor="hand2",
                                    command=self._analiz_calistir)
        self.analiz_btn.pack(side="left")

        a_info = tk.Frame(tab_analiz, bg="#EDE7F6", bd=1, relief="solid")
        a_info.pack(fill="x", padx=20, pady=(0, 8))
        tk.Label(a_info,
                 text="Bu sayfa Excel’deki Bilanço, Gelir Tablosu ve Nakit Akım sayfalarını okur; "
                      "oranları hesaplar ve yerleşik AI metnini üretir. İsteğe bağlı olarak "
                      "OPENAI_API_KEY ortam değişkeni varsa kısa yönetici özeti eklenir.",
                 font=("Arial", 8), fg="#4527A0", bg="#EDE7F6", justify="left", wraplength=640).pack(anchor="w", padx=8, pady=6)

        self.analiz_pb = ttk.Progressbar(tab_analiz, mode="indeterminate", length=500)
        self.analiz_pb.pack(pady=(0, 6))

        self.analiz_text = stext.ScrolledText(tab_analiz, font=("Consolas", 9), height=22,
                                              bg="#1A1A2E", fg="#E8E8E8", bd=1, relief="flat",
                                              wrap="word", padx=10, pady=8)
        self.analiz_text.pack(fill="both", expand=True, padx=20, pady=(0, 8))
        self.analiz_text.insert("1.0", "Excel dosyasını seçin veya önce «Veri çek & Excel» sekmesinde dosya oluşturun; "
                                        "oluşan dosya yolu burada otomatik dolar.\n")

    def _klasor_sec(self):
        k = filedialog.askdirectory(title="Kayıt klasörünü seç",
                                    initialdir=self.klasor_var.get())
        if k: self.klasor_var.set(k)

    def _analiz_dosya_sec(self):
        cur = self.analiz_dosya_var.get().strip()
        init = os.path.dirname(cur) if cur and os.path.isfile(cur) else (self.klasor_var.get() or ".")
        p = filedialog.askopenfilename(
            title="Finansal tablolar Excel dosyası",
            filetypes=[("Excel", "*.xlsx"), ("Tümü", "*.*")],
            initialdir=init,
        )
        if p:
            self.analiz_dosya_var.set(p)

    def _analiz_calistir(self):
        path = self.analiz_dosya_var.get().strip()
        if not path or not os.path.isfile(path):
            messagebox.showwarning("Uyarı", "Geçerli bir .xlsx dosyası seçin.")
            return
        self.analiz_btn.configure(state="disabled", text="⏳  Analiz ediliyor...")
        self.analiz_pb.start(12)
        llm = self.analiz_llm_var.get()
        threading.Thread(target=self._analiz_thread, args=(path, llm), daemon=True).start()

    def _analiz_thread(self, path, llm_istek):
        def silent_log(_msg):
            pass
        try:
            tablolar = excel_den_tablolar_oku(path, silent_log)
            if not tablolar.get("Bilanço"):
                self.after(0, lambda: self._analiz_ui_done_error(
                    "Bilanço sayfası bulunamadı veya okunamadı.\nBu uygulamanın ürettiği .xlsx dosyasını kullanın."))
                return
            ye, yn = analiz_donemleri_bul(tablolar)
            if not ye or not yn:
                self.after(0, lambda: self._analiz_ui_done_error(
                    "Dönem sütunları (ör. 2024, 2025) okunamadı."))
                return
            hisse = hisse_kodu_excel_yolundan(path)
            metin = analiz_metni_olustur(tablolar, ye, yn, hisse, silent_log, llm_istek=llm_istek)
            self.after(0, lambda m=metin: self._analiz_ui_done_ok(m))
        except Exception as e:
            err = f"{e}\n\n{traceback.format_exc()}"
            self.after(0, lambda msg=err: self._analiz_ui_done_error(msg))

    def _analiz_ui_done_ok(self, metin):
        self.analiz_pb.stop()
        self.analiz_btn.configure(state="normal", text="▶  Analizi çalıştır")
        self.analiz_text.delete("1.0", "end")
        self.analiz_text.insert("1.0", metin)

    def _analiz_ui_done_error(self, msg):
        try:
            self.analiz_pb.stop()
        except tk.TclError:
            pass
        self.analiz_btn.configure(state="normal", text="▶  Analizi çalıştır")
        messagebox.showerror("Analiz hatası", str(msg)[:800])

    def _log(self, mesaj):
        """Log alanına thread-safe mesaj ekle."""
        def _yaz():
            self.log_alan.configure(state="normal")
            self.log_alan.insert("end", mesaj + "\n")
            self.log_alan.see("end")
            self.log_alan.configure(state="disabled")
        self.after(0, _yaz)

    def _baslat(self):
        hisse = self.hisse_var.get().strip().upper()
        if not hisse:
            messagebox.showwarning("Uyarı", "Lütfen bir hisse kodu girin.")
            return
        try:
            baslangic = int(self.baslangic_var.get().strip())
            bitis = int(self.bitis_var.get().strip())
        except Exception:
            messagebox.showwarning("Uyarı", "Lütfen geçerli yıl seçin.")
            return
        if baslangic > bitis:
            messagebox.showwarning("Uyarı", "Başlangıç yılı, bitiş yılından büyük olamaz.")
            return
        donem_lbl = self.donem_label_var.get().strip()
        donem_ay = self.donem_map.get(donem_lbl, "12")

        # Sonuç alanını temizle
        for w in self.sonuc_frame.winfo_children(): w.destroy()

        # Log temizle
        self.log_alan.configure(state="normal")
        self.log_alan.delete("1.0","end")
        self.log_alan.configure(state="disabled")

        # UI'yi kilitle
        self.btn.configure(state="disabled", text="⏳  İşleniyor...")
        self.pb.start(12)
        self.donem_ust_etiket.configure(text=f"{baslangic} → {bitis} | {donem_ay}")

        kayit = self.klasor_var.get() or os.path.expanduser("~")

        threading.Thread(
            target=isle,
            args=(hisse, kayit, self._log, self._bitti, self._hata, baslangic, bitis, donem_ay),
            daemon=True
        ).start()

    def _bitti(self, dosya_yolu):
        self._son_dosya = dosya_yolu
        self.after(0, lambda: self._sonuc_goster(dosya_yolu))

    def _sonuc_goster(self, dosya_yolu):
        self.pb.stop(); self.pb.configure(value=0)
        self.btn.configure(state="normal", text="▶  Excel Oluştur")

        # Yeşil başarı kutusu
        kutu = tk.Frame(self.sonuc_frame, bg="#E6F4EA", bd=1, relief="solid")
        kutu.pack(fill="x", pady=4)

        tk.Label(kutu, text="✅  Excel başarıyla oluşturuldu!",
                 font=("Arial",11,"bold"), fg="#2E7D32", bg="#E6F4EA",
                 pady=8).pack(anchor="w", padx=12)

        # Dosya yolu
        yol_frame = tk.Frame(kutu, bg="#E6F4EA")
        yol_frame.pack(fill="x", padx=12, pady=(0,6))
        tk.Label(yol_frame, text="📁 Konum:", font=("Arial",9),
                 fg="#555", bg="#E6F4EA").pack(side="left")
        tk.Label(yol_frame, text=dosya_yolu, font=("Arial",9),
                 fg="#1A6B3C", bg="#E6F4EA", cursor="hand2").pack(side="left", padx=4)

        # Butonlar
        btn_frame = tk.Frame(kutu, bg="#E6F4EA")
        btn_frame.pack(fill="x", padx=12, pady=(0,8))

        tk.Button(btn_frame, text="📂  Excel'i Aç",
                  font=("Arial",10,"bold"), bg="#1F4E79", fg="white",
                  activebackground="#163A5F", bd=0, padx=14, pady=6,
                  cursor="hand2",
                  command=lambda: os.startfile(dosya_yolu)).pack(side="left", padx=(0,8))

        tk.Button(btn_frame, text="📁  Klasörü Aç",
                  font=("Arial",10), bg="#E8F5E9", fg="#2E7D32",
                  activebackground="#C8E6C9", bd=1, relief="groove",
                  padx=14, pady=6, cursor="hand2",
                  command=lambda: os.startfile(os.path.dirname(dosya_yolu))).pack(side="left")

        try:
            self.analiz_dosya_var.set(dosya_yolu)
        except tk.TclError:
            pass

    def _hata(self, mesaj):
        self.after(0, lambda: self._hata_goster(mesaj))

    def _hata_goster(self, mesaj):
        self.pb.stop()
        self.btn.configure(state="normal", text="▶  Excel Oluştur")

        kutu = tk.Frame(self.sonuc_frame, bg="#FEEBEB", bd=1, relief="solid")
        kutu.pack(fill="x", pady=4)
        tk.Label(kutu, text="❌  Hata oluştu", font=("Arial",11,"bold"),
                 fg="#C62828", bg="#FEEBEB", pady=8).pack(anchor="w", padx=12)
        tk.Label(kutu, text=str(mesaj)[:200], font=("Arial",9),
                 fg="#555", bg="#FEEBEB", wraplength=560, justify="left",
                 pady=4).pack(anchor="w", padx=12)

# ══════════════════════════════════════════════════════════════════════════════
# 6.  Başlat
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    app = App()
    app.mainloop()